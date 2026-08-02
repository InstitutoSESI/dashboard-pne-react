"""Parser auditavel da Sinopse municipal da Educacao Superior.

O modulo implementa somente o escopo ESUP-1. Ele le sete tabelas anuais,
normaliza os fatos aprovados e produz evidencias temporarias de qualidade.
Nao grava banco e nao escreve em ``public/data``.
"""

from __future__ import annotations

import hashlib
import json
import math
import re
import tempfile
import unicodedata
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Iterable, Mapping, Sequence

from openpyxl import load_workbook

from .municipality_registry import MunicipalityRegistry
from .state_config import StateConfig


SUPPORTED_YEARS = tuple(range(2018, 2025))
REQUIRED_TABLES = ("1.1", "2.1", "2.3", "5.1", "7.1", "7.2", "7.3")

STATUS_OBSERVED = "observed"
STATUS_DERIVED_ZERO = "derived_zero"
STATUS_UNAVAILABLE = "unavailable"
STATUS_NOT_APPLICABLE = "not_applicable"
DATA_STATUSES = {
    STATUS_OBSERVED,
    STATUS_DERIVED_ZERO,
    STATUS_UNAVAILABLE,
    STATUS_NOT_APPLICABLE,
}

UNIVERSE_GRADUATION = "graduation"
UNIVERSE_PRESENTIAL_GRADUATION = "presential_graduation"
UNIVERSE_DISTANCE_GRADUATION = "distance_graduation"
UNIVERSE_IES = "institutions_offering_graduation_or_sequential"
UNIVERSE_FACULTY = "faculty_in_graduation_or_sequential"

TERRITORY_OFFER = "course_offer_location"
TERRITORY_EAD = "ead_offer_location"
TERRITORY_IES = "ies_administrative_headquarters"
TERRITORY_FACULTY = "faculty_institution_headquarters"

DIRECT_METRICS = (
    "enrollments_total",
    "enrollments_presential",
    "enrollments_distance",
    "ies_headquarters",
    "ead_poles",
    "vacancies_presential",
    "entrants_total",
    "graduates_total",
    "faculty_total",
)

DECOMPOSITION_METRICS = (
    "enrollments_by_dependency",
    "enrollments_by_organization",
    "ies_by_dependency",
    "ies_by_organization",
    "faculty_by_education",
)

DEPENDENCIES = ("federal", "estadual", "municipal", "privada")
ORGANIZATIONS = (
    "universidade",
    "centro_universitario",
    "faculdade",
    "instituto_federal_cefet",
)
FACULTY_EDUCATION = (
    "Sem Graduação",
    "Graduação",
    "Especialização",
    "Mestrado",
    "Doutorado",
)

TITLE_REQUIREMENTS = {
    "1.1": ("numero de instituicoes de ensino superior", "sede administrativa"),
    "2.1": ("numero de docentes em exercicio", "municipio"),
    "2.3": ("numero de docentes em exercicio", "grau de formacao"),
    "5.1": (
        "numero de matriculas",
        "cursos de graduacao e sequenciais",
        "presenciais e a distancia",
    ),
    "7.1": ("numero de ingressantes", "matriculas", "concluintes"),
    "7.2": ("cursos de graduacao", "presenciais"),
    "7.3": ("cursos de graduacao a distancia", "numero de polos"),
}


@dataclass(frozen=True)
class ColumnRule:
    name: str
    includes: tuple[str, ...]
    leaf: str
    part_count: int | None = None
    excludes: tuple[str, ...] = ()


@dataclass(frozen=True)
class MetricRule:
    metric: str
    column: ColumnRule
    universe: str
    territorial_reference: str
    modality: str | None = None
    administrative_dependency: str | None = None
    academic_organization: str | None = None
    faculty_education: str | None = None
    public_record: bool = True


@dataclass
class SheetLayout:
    year: int
    table: str
    sheet_name: str
    title: str
    header_start_row: int
    header_end_row: int
    first_source_row: int
    dimensions: dict[str, int]
    semantic_columns: dict[str, int]
    semantic_header_paths: dict[str, str]


@dataclass
class NormalizedRecord:
    year: int
    municipality_id: str
    municipality: str
    metric: str
    value: int | None
    modality: str | None
    administrative_dependency: str | None
    academic_organization: str | None
    faculty_education: str | None
    statistical_universe: str
    territorial_reference: str
    status: str
    source_file: str
    source_table: str
    semantic_header_path: str
    source_sha256: str

    def key(self) -> tuple:
        return (
            self.year,
            self.municipality_id,
            self.metric,
            self.modality,
            self.administrative_dependency,
            self.academic_organization,
            self.faculty_education,
        )


@dataclass
class ParsedSheet:
    records: list[NormalizedRecord]
    controls: dict[str, dict[str, int | None]]
    state_controls: dict[str, int | None]
    layout: SheetLayout
    municipality_count: int


@dataclass
class ParsedAudit:
    records: list[NormalizedRecord]
    quality_report: dict
    pilots: list[dict]
    source_files: list[dict]


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def normalize_header(value: object) -> str:
    normalized = unicodedata.normalize("NFKD", normalize_text(value))
    normalized = "".join(
        character for character in normalized if not unicodedata.combining(character)
    )
    return " ".join(
        "".join(
            character if character.isalnum() else " "
            for character in normalized.lower()
        ).split()
    )


def extract_year(path: Path) -> int:
    matches = {
        int(match)
        for match in re.findall(r"(?<!\d)(20\d{2})(?!\d)", path.stem)
    }
    if len(matches) != 1:
        raise ValueError(
            f"Nao foi possivel identificar um unico ano no arquivo {path.name!r}."
        )
    return matches.pop()


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def discover_source_files(
    source_dir: Path,
    years: Iterable[int] | None = None,
) -> dict[int, Path]:
    source_dir = source_dir.resolve()
    if not source_dir.is_dir():
        raise ValueError(f"Diretorio de fontes inexistente: {source_dir}.")

    requested_years = set(years or SUPPORTED_YEARS)
    discovered: dict[int, Path] = {}
    for path in sorted(source_dir.glob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        year = extract_year(path)
        if year not in requested_years:
            continue
        if year in discovered:
            raise ValueError(
                f"Mais de um arquivo encontrado para {year}: "
                f"{discovered[year].name!r} e {path.name!r}."
            )
        discovered[year] = path

    missing = sorted(requested_years - set(discovered))
    if missing:
        raise ValueError(f"Arquivos anuais ausentes: {missing}.")
    return dict(sorted(discovered.items()))


def municipality_universe_from_registry(
    registry: MunicipalityRegistry,
) -> dict[str, str]:
    return {
        record.ibge_code: record.name
        for record in registry.ordered_records
    }


def normalize_ibge_code(value: object, *, field: str) -> str:
    if value is None or isinstance(value, bool):
        raise ValueError(f"Codigo IBGE invalido em {field}: {value!r}.")
    code = normalize_text(value)
    if re.fullmatch(r"\d{7}", code) is None:
        raise ValueError(f"Codigo IBGE deve ter sete digitos em {field}: {value!r}.")
    return code


def optional_non_negative_integer(
    value: object,
    *,
    field: str,
    row_number: int,
) -> int | None:
    if value is None or normalize_text(value) in {"", "-", "—", "..."}:
        return None
    if isinstance(value, bool):
        raise ValueError(
            f"Valor invalido em {field}, linha {row_number}: {value!r}."
        )
    try:
        numeric = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(
            f"Valor nao numerico em {field}, linha {row_number}: {value!r}."
        ) from exc
    if not math.isfinite(numeric) or not numeric.is_integer():
        raise ValueError(
            f"Valor nao inteiro em {field}, linha {row_number}: {value!r}."
        )
    integer = int(numeric)
    if integer < 0:
        raise ValueError(f"Valor negativo em {field}, linha {row_number}.")
    return integer


def _sheet_names(workbook) -> dict[str, str]:
    found: dict[str, str] = {}
    for sheet_name in workbook.sheetnames:
        match = re.match(r"^(\d+\.\d+)", sheet_name)
        if not match:
            continue
        table = match.group(1)
        if table not in REQUIRED_TABLES:
            continue
        if table in found:
            raise ValueError(
                f"Mais de uma aba encontrada para a tabela {table}: "
                f"{found[table]!r} e {sheet_name!r}."
            )
        found[table] = sheet_name
    missing = sorted(set(REQUIRED_TABLES) - set(found))
    if missing:
        raise ValueError(f"Abas obrigatorias ausentes: {missing}.")
    return found


def _sheet_title(sheet) -> str:
    values: list[str] = []
    for row in sheet.iter_rows(min_row=1, max_row=5, values_only=True):
        for value in row:
            text_value = normalize_text(value)
            if text_value and text_value not in values:
                values.append(text_value)
    if not values:
        raise ValueError(f"Titulo ausente na aba {sheet.title}.")
    return max(values, key=len)


def _validate_title(table: str, title: str, year: int) -> None:
    normalized = normalize_header(title)
    for token in TITLE_REQUIREMENTS[table]:
        if token not in normalized:
            raise ValueError(
                f"Titulo inesperado na tabela {table} de {year}: "
                f"marcador {token!r} ausente."
            )
    if str(year) not in title:
        raise ValueError(f"O titulo da tabela {table} nao identifica o ano {year}.")


def _header_bounds(sheet) -> tuple[int, int, int]:
    header_start = None
    first_source_row = None
    for row_number, row in enumerate(
        sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 30), values_only=True),
        start=1,
    ):
        normalized = [normalize_header(value) for value in row]
        if (
            header_start is None
            and "regiao geografica" in normalized
            and "unidade da federacao" in normalized
        ):
            header_start = row_number
        if any(normalize_text(value) == "Total Brasil" for value in row):
            first_source_row = row_number
            break
    if header_start is None or first_source_row is None:
        raise ValueError(f"Cabecalho nao reconhecido na aba {sheet.title}.")
    return header_start, first_source_row - 1, first_source_row


def _header_paths(
    sheet,
    header_start: int,
    header_end: int,
) -> tuple[dict[int, tuple[str, ...]], dict[int, tuple[str, ...]]]:
    rows = list(
        sheet.iter_rows(
            min_row=header_start,
            max_row=header_end,
            values_only=True,
        )
    )
    last_column = 0
    for row in rows:
        for index, value in enumerate(row, start=1):
            if normalize_text(value):
                last_column = max(last_column, index)
    if last_column == 0:
        raise ValueError(f"Cabecalho vazio na aba {sheet.title}.")

    display_by_row: list[list[str]] = []
    normalized_by_row: list[list[str]] = []
    for row_index, row in enumerate(rows):
        display_values: list[str] = []
        normalized_values: list[str] = []
        current_display = ""
        current_normalized = ""
        for index in range(last_column):
            raw = row[index] if index < len(row) else None
            if normalize_text(raw):
                current_display = normalize_text(raw)
                current_normalized = normalize_header(raw)
            elif any(
                normalize_text(previous[index] if index < len(previous) else None)
                for previous in rows[:row_index]
            ):
                # Uma nova celula-pai na mesma coluna encerra o preenchimento
                # horizontal herdado do grupo anterior. Isso distingue, por
                # exemplo, o Total de Universidade do ultimo subtipo de
                # dependencia administrativa imediatamente a esquerda.
                current_display = ""
                current_normalized = ""
            display_values.append(current_display)
            normalized_values.append(current_normalized)
        display_by_row.append(display_values)
        normalized_by_row.append(normalized_values)

    display_paths: dict[int, tuple[str, ...]] = {}
    normalized_paths: dict[int, tuple[str, ...]] = {}
    for column in range(1, last_column + 1):
        display_parts: list[str] = []
        normalized_parts: list[str] = []
        for row_index in range(len(rows)):
            display_value = display_by_row[row_index][column - 1]
            normalized_value = normalized_by_row[row_index][column - 1]
            if display_value and (
                not display_parts or display_value != display_parts[-1]
            ):
                display_parts.append(display_value)
            if normalized_value and (
                not normalized_parts or normalized_value != normalized_parts[-1]
            ):
                normalized_parts.append(normalized_value)
        display_paths[column] = tuple(display_parts)
        normalized_paths[column] = tuple(normalized_parts)
    return display_paths, normalized_paths


def _dimension_columns(
    normalized_paths: Mapping[int, tuple[str, ...]],
) -> dict[str, int]:
    expected = {
        "uf": "unidade da federacao",
        "municipality": "municipio",
        "municipality_id": "codigo do municipio",
        "academic_level": "nivel academico do curso",
    }
    dimensions: dict[str, int] = {}
    for key, label in expected.items():
        candidates = [
            column
            for column, parts in normalized_paths.items()
            if parts and parts[0] == label
        ]
        if len(candidates) > 1:
            raise ValueError(f"Dimensao ambigua {label!r}: {candidates}.")
        if candidates:
            dimensions[key] = candidates[0]
    required = {"uf", "municipality", "municipality_id"}
    if not required.issubset(dimensions):
        raise ValueError(f"Dimensoes territoriais ausentes: {dimensions}.")
    return dimensions


def _column_matches(parts: tuple[str, ...], rule: ColumnRule) -> bool:
    if rule.part_count is not None and len(parts) != rule.part_count:
        return False
    if not parts or rule.leaf != parts[-1]:
        return False
    joined = " | ".join(parts)
    if any(token not in joined for token in rule.includes):
        return False
    if any(token in joined for token in rule.excludes):
        return False
    return True


def _select_column(
    normalized_paths: Mapping[int, tuple[str, ...]],
    rule: ColumnRule,
) -> int:
    candidates = [
        column
        for column, parts in normalized_paths.items()
        if _column_matches(parts, rule)
    ]
    if len(candidates) != 1:
        raise ValueError(
            f"Coluna semantica {rule.name!r} ambigua ou ausente: {candidates}."
        )
    return candidates[0]


def _overall_total_rule(group: str) -> ColumnRule:
    return ColumnRule(
        name="total",
        includes=(group,),
        leaf="total",
        part_count=2,
    )


def _dependency_rule(name: str) -> ColumnRule:
    if name == "privada":
        return ColumnRule(
            name=f"dependency.{name}",
            includes=("total por dependencia administrativa", "privada"),
            leaf="total",
            part_count=4,
        )
    return ColumnRule(
        name=f"dependency.{name}",
        includes=("total por dependencia administrativa", "publica"),
        leaf=name,
        part_count=4,
    )


def _organization_rule(name: str) -> ColumnRule:
    labels = {
        "universidade": "universidade",
        "centro_universitario": "centro universitario",
        "faculdade": "faculdade",
        "instituto_federal_cefet": "instituto federal de educacao",
    }
    return ColumnRule(
        name=f"organization.{name}",
        includes=(labels[name],),
        leaf="federal" if name == "instituto_federal_cefet" else "total",
        part_count=None if name == "instituto_federal_cefet" else 3,
    )


def _metric_rules(table: str) -> tuple[MetricRule, ...]:
    organization_group = "organizacao academica e dependencia administrativa"
    if table == "1.1":
        return (
            MetricRule(
                "ies_headquarters",
                _overall_total_rule(organization_group),
                UNIVERSE_IES,
                TERRITORY_IES,
            ),
            *(
                MetricRule(
                    "ies_by_dependency",
                    _dependency_rule(dependency),
                    UNIVERSE_IES,
                    TERRITORY_IES,
                    administrative_dependency=dependency,
                )
                for dependency in DEPENDENCIES
            ),
            *(
                MetricRule(
                    "ies_by_organization",
                    _organization_rule(organization),
                    UNIVERSE_IES,
                    TERRITORY_IES,
                    academic_organization=organization,
                )
                for organization in ORGANIZATIONS
            ),
        )
    if table == "2.1":
        return (
            MetricRule(
                "faculty_total",
                _overall_total_rule(organization_group),
                UNIVERSE_FACULTY,
                TERRITORY_FACULTY,
            ),
        )
    if table == "2.3":
        formation_group = "total por sexo e por grau de formacao"
        rules = [
            MetricRule(
                "faculty_total_control",
                _overall_total_rule(
                    "organizacao academica dependencia administrativa sexo e grau"
                ),
                UNIVERSE_FACULTY,
                TERRITORY_FACULTY,
                public_record=False,
            )
        ]
        for display_name in FACULTY_EDUCATION:
            normalized_name = normalize_header(display_name)
            rules.append(
                MetricRule(
                    "faculty_by_education",
                    ColumnRule(
                        name=f"faculty_education.{normalized_name}",
                        includes=(formation_group, "grau de formacao"),
                        leaf=normalized_name,
                        part_count=4,
                        excludes=(
                            ("sem graduacao",)
                            if normalized_name == "graduacao"
                            else ()
                        ),
                    ),
                    UNIVERSE_FACULTY,
                    TERRITORY_FACULTY,
                    faculty_education=display_name,
                )
            )
        return tuple(rules)
    if table == "5.1":
        rules = [
            MetricRule(
                "enrollments_total_control",
                _overall_total_rule(organization_group),
                UNIVERSE_GRADUATION,
                TERRITORY_OFFER,
                public_record=False,
            )
        ]
        rules.extend(
            MetricRule(
                "enrollments_by_dependency",
                _dependency_rule(dependency),
                UNIVERSE_GRADUATION,
                TERRITORY_OFFER,
                administrative_dependency=dependency,
            )
            for dependency in DEPENDENCIES
        )
        rules.extend(
            MetricRule(
                "enrollments_by_organization",
                _organization_rule(organization),
                UNIVERSE_GRADUATION,
                TERRITORY_OFFER,
                academic_organization=organization,
            )
            for organization in ORGANIZATIONS
        )
        return tuple(rules)
    if table == "7.1":
        return (
            MetricRule(
                "entrants_total",
                ColumnRule("entrants_total", ("total geral",), "ingressantes", 2),
                UNIVERSE_GRADUATION,
                TERRITORY_OFFER,
            ),
            MetricRule(
                "enrollments_total",
                ColumnRule("enrollments_total", ("total geral",), "matriculas", 2),
                UNIVERSE_GRADUATION,
                TERRITORY_OFFER,
            ),
            MetricRule(
                "graduates_total",
                ColumnRule("graduates_total", ("total geral",), "concluintes", 2),
                UNIVERSE_GRADUATION,
                TERRITORY_OFFER,
            ),
        )
    if table == "7.2":
        return (
            MetricRule(
                "vacancies_presential",
                ColumnRule(
                    "vacancies_presential",
                    ("total geral",),
                    "vagas oferecidas",
                    2,
                ),
                UNIVERSE_PRESENTIAL_GRADUATION,
                TERRITORY_OFFER,
                modality="presential",
            ),
            MetricRule(
                "enrollments_presential",
                ColumnRule(
                    "enrollments_presential",
                    ("total geral",),
                    "matriculas",
                    2,
                ),
                UNIVERSE_PRESENTIAL_GRADUATION,
                TERRITORY_OFFER,
                modality="presential",
            ),
        )
    if table == "7.3":
        return (
            MetricRule(
                "ead_poles",
                ColumnRule("ead_poles", ("total geral",), "numero de polos", 2),
                UNIVERSE_DISTANCE_GRADUATION,
                TERRITORY_EAD,
                modality="distance",
            ),
            MetricRule(
                "enrollments_distance",
                ColumnRule(
                    "enrollments_distance",
                    ("total geral",),
                    "matriculas",
                    2,
                ),
                UNIVERSE_DISTANCE_GRADUATION,
                TERRITORY_EAD,
                modality="distance",
            ),
        )
    raise ValueError(f"Tabela sem regras: {table}.")


def _row_value(
    row: Sequence[object],
    column: int,
    *,
    field_name: str,
    row_number: int,
) -> int | None:
    raw = row[column - 1] if column <= len(row) else None
    return optional_non_negative_integer(
        raw,
        field=field_name,
        row_number=row_number,
    )


def _parse_sheet(
    workbook,
    *,
    year: int,
    table: str,
    sheet_name: str,
    source_file: Path,
    source_sha256: str,
    state_config: StateConfig,
    municipality_universe: Mapping[str, str],
) -> ParsedSheet:
    sheet = workbook[sheet_name]
    title = _sheet_title(sheet)
    _validate_title(table, title, year)
    header_start, header_end, first_source_row = _header_bounds(sheet)
    display_paths, normalized_paths = _header_paths(sheet, header_start, header_end)
    dimensions = _dimension_columns(normalized_paths)
    if table in {"5.1", "7.1"} and "academic_level" not in dimensions:
        raise ValueError(f"A tabela {table} de {year} nao possui nivel academico.")

    rules = _metric_rules(table)
    metric_columns = {
        rule.column.name: _select_column(normalized_paths, rule.column)
        for rule in rules
    }
    semantic_paths = {
        rule.column.name: " > ".join(display_paths[metric_columns[rule.column.name]])
        for rule in rules
    }

    records: list[NormalizedRecord] = []
    controls: dict[str, dict[str, int | None]] = {}
    state_controls: dict[str, int | None] = {}
    seen_municipalities: set[str] = set()
    target_state_name = normalize_header(state_config.state_name)

    for row_number, row in enumerate(
        sheet.iter_rows(min_row=first_source_row, values_only=True),
        start=first_source_row,
    ):
        uf_value = normalize_header(row[dimensions["uf"] - 1])
        academic_level = (
            normalize_header(row[dimensions["academic_level"] - 1])
            if "academic_level" in dimensions
            else ""
        )
        if table in {"5.1", "7.1"} and academic_level != "graduacao":
            continue

        is_state_total = uf_value == f"total {target_state_name}"
        is_target_municipality = uf_value == target_state_name
        if not is_state_total and not is_target_municipality:
            continue

        values_by_rule: dict[str, int | None] = {}
        for rule in rules:
            column = metric_columns[rule.column.name]
            values_by_rule[rule.column.name] = _row_value(
                row,
                column,
                field_name=f"{table}.{rule.column.name}",
                row_number=row_number,
            )

        if is_state_total:
            for rule in rules:
                state_controls[rule.metric] = values_by_rule[rule.column.name]
            continue

        municipality_name = normalize_text(row[dimensions["municipality"] - 1])
        raw_code = row[dimensions["municipality_id"] - 1]
        if not municipality_name and raw_code is None:
            continue
        municipality_id = normalize_ibge_code(
            raw_code,
            field=f"{table}, linha {row_number}",
        )
        if municipality_id not in municipality_universe:
            raise ValueError(
                f"Codigo {municipality_id} da tabela {table} nao pertence "
                "ao universo municipal do projeto."
            )
        if municipality_id in seen_municipalities:
            raise ValueError(
                f"Municipio duplicado na tabela {table} de {year}: "
                f"{municipality_id}."
            )
        seen_municipalities.add(municipality_id)

        official_name = municipality_universe[municipality_id]
        if municipality_name != official_name:
            raise ValueError(
                f"Nome municipal divergente na tabela {table} de {year}: "
                f"{municipality_id} usa {municipality_name!r}, esperado "
                f"{official_name!r}."
            )
        controls[municipality_id] = {}
        for rule in rules:
            value = values_by_rule[rule.column.name]
            controls[municipality_id][rule.metric] = value
            if not rule.public_record:
                continue
            records.append(
                NormalizedRecord(
                    year=year,
                    municipality_id=municipality_id,
                    municipality=official_name,
                    metric=rule.metric,
                    value=value,
                    modality=rule.modality,
                    administrative_dependency=rule.administrative_dependency,
                    academic_organization=rule.academic_organization,
                    faculty_education=rule.faculty_education,
                    statistical_universe=rule.universe,
                    territorial_reference=rule.territorial_reference,
                    status=(
                        STATUS_OBSERVED if value is not None else STATUS_UNAVAILABLE
                    ),
                    source_file=source_file.name,
                    source_table=table,
                    semantic_header_path=semantic_paths[rule.column.name],
                    source_sha256=source_sha256,
                )
            )

    layout = SheetLayout(
        year=year,
        table=table,
        sheet_name=sheet_name,
        title=title,
        header_start_row=header_start,
        header_end_row=header_end,
        first_source_row=first_source_row,
        dimensions=dimensions,
        semantic_columns=metric_columns,
        semantic_header_paths=semantic_paths,
    )
    return ParsedSheet(
        records=records,
        controls=controls,
        state_controls=state_controls,
        layout=layout,
        municipality_count=len(seen_municipalities),
    )


def _records_index(records: Sequence[NormalizedRecord]) -> dict[tuple, NormalizedRecord]:
    index: dict[tuple, NormalizedRecord] = {}
    for record in records:
        key = record.key()
        if key in index:
            raise ValueError(f"Chave normalizada duplicada: {key}.")
        index[key] = record
    return index


def _direct_index(
    records: Sequence[NormalizedRecord],
) -> dict[tuple[int, str, str], NormalizedRecord]:
    index: dict[tuple[int, str, str], NormalizedRecord] = {}
    for record in records:
        if record.metric not in DIRECT_METRICS:
            continue
        key = (record.year, record.municipality_id, record.metric)
        if key in index:
            raise ValueError(f"Indicador direto duplicado: {key}.")
        index[key] = record
    return index


def _derived_zero_record(
    *,
    total: NormalizedRecord,
    missing_metric: str,
    source_file: dict,
) -> NormalizedRecord:
    is_presential = missing_metric == "enrollments_presential"
    return NormalizedRecord(
        year=total.year,
        municipality_id=total.municipality_id,
        municipality=total.municipality,
        metric=missing_metric,
        value=0,
        modality="presential" if is_presential else "distance",
        administrative_dependency=None,
        academic_organization=None,
        faculty_education=None,
        statistical_universe=(
            UNIVERSE_PRESENTIAL_GRADUATION
            if is_presential
            else UNIVERSE_DISTANCE_GRADUATION
        ),
        territorial_reference=TERRITORY_OFFER if is_presential else TERRITORY_EAD,
        status=STATUS_DERIVED_ZERO,
        source_file=source_file["fileName"],
        source_table="7.2" if is_presential else "7.3",
        semantic_header_path=(
            "Derivado: matriculas totais observadas = outra modalidade observada"
        ),
        source_sha256=source_file["sha256"],
    )


def apply_derived_zero_rule(
    records: list[NormalizedRecord],
    source_files: Sequence[dict],
) -> list[dict]:
    by_year_file = {item["year"]: item for item in source_files}
    direct = _direct_index(records)
    existing_keys = set(_records_index(records))
    derived: list[dict] = []
    totals = [
        record
        for record in records
        if record.metric == "enrollments_total"
        and record.status == STATUS_OBSERVED
        and record.value is not None
    ]
    for total in totals:
        base = (total.year, total.municipality_id)
        presential = direct.get((*base, "enrollments_presential"))
        distance = direct.get((*base, "enrollments_distance"))
        presential_observed = (
            presential is not None
            and presential.status == STATUS_OBSERVED
            and presential.value is not None
        )
        distance_observed = (
            distance is not None
            and distance.status == STATUS_OBSERVED
            and distance.value is not None
        )
        if (
            not presential_observed
            and distance_observed
            and distance.value == total.value
        ):
            record = _derived_zero_record(
                total=total,
                missing_metric="enrollments_presential",
                source_file=by_year_file[total.year],
            )
        elif (
            not distance_observed
            and presential_observed
            and presential.value == total.value
        ):
            record = _derived_zero_record(
                total=total,
                missing_metric="enrollments_distance",
                source_file=by_year_file[total.year],
            )
        else:
            continue
        if record.key() not in existing_keys:
            records.append(record)
            existing_keys.add(record.key())
            direct[(record.year, record.municipality_id, record.metric)] = record
            derived.append(
                {
                    "year": record.year,
                    "municipalityId": record.municipality_id,
                    "municipality": record.municipality,
                    "metric": record.metric,
                    "status": record.status,
                }
            )
    return derived


def _severity(relative_difference: float | None, absolute_difference: int) -> str:
    if absolute_difference == 0:
        return "none"
    if relative_difference is None or relative_difference > 0.01:
        return "high"
    if relative_difference > 0.001:
        return "medium"
    return "low"


def _comparison(
    *,
    year: int,
    municipality_id: str,
    municipality: str,
    measure: str,
    left_label: str,
    left_value: int | None,
    right_label: str,
    right_value: int | None,
    possible_cause: str,
    scope: str = "municipality",
) -> dict:
    if left_value is None or right_value is None:
        return {
            "year": year,
            "municipalityId": municipality_id,
            "municipality": municipality,
            "scope": scope,
            "measure": measure,
            "status": "unavailable",
            "left": {"label": left_label, "value": left_value},
            "right": {"label": right_label, "value": right_value},
            "absoluteDifference": None,
            "relativeDifference": None,
            "possibleCause": "Uma das medidas comparadas esta ausente.",
            "severity": "medium",
        }
    absolute = left_value - right_value
    relative = abs(absolute) / abs(left_value) if left_value else None
    return {
        "year": year,
        "municipalityId": municipality_id,
        "municipality": municipality,
        "scope": scope,
        "measure": measure,
        "status": "matched" if absolute == 0 else "mismatched",
        "left": {"label": left_label, "value": left_value},
        "right": {"label": right_label, "value": right_value},
        "absoluteDifference": absolute,
        "relativeDifference": relative,
        "possibleCause": None if absolute == 0 else possible_cause,
        "severity": _severity(relative, absolute),
    }


def _sum_records(
    records: Sequence[NormalizedRecord],
    *,
    year: int,
    municipality_id: str,
    metric: str,
) -> int | None:
    selected = [
        record
        for record in records
        if record.year == year
        and record.municipality_id == municipality_id
        and record.metric == metric
    ]
    if not selected or any(record.value is None for record in selected):
        return None
    return sum(int(record.value) for record in selected if record.value is not None)


def build_reconciliations(
    records: Sequence[NormalizedRecord],
    controls: Mapping[tuple[int, str, str], Mapping[str, int | None]],
    state_controls: Mapping[tuple[int, str], Mapping[str, int | None]],
    municipality_universe: Mapping[str, str],
    state_config: StateConfig,
) -> tuple[list[dict], list[dict], list[dict]]:
    direct = _direct_index(records)
    years = sorted({record.year for record in records})
    comparisons: list[dict] = []
    formation_exhaustiveness: list[dict] = []
    modality_unavailable: list[dict] = []

    def direct_value(year: int, municipality_id: str, metric: str) -> int | None:
        record = direct.get((year, municipality_id, metric))
        return record.value if record is not None else None

    for year in years:
        municipality_ids = sorted(
            {
                record.municipality_id
                for record in records
                if record.year == year
            }
        )
        for municipality_id in municipality_ids:
            municipality = municipality_universe[municipality_id]
            total_71 = direct_value(year, municipality_id, "enrollments_total")
            total_51 = controls.get((year, "5.1", municipality_id), {}).get(
                "enrollments_total_control"
            )
            comparisons.append(
                _comparison(
                    year=year,
                    municipality_id=municipality_id,
                    municipality=municipality,
                    measure="enrollments_7.1_vs_5.1",
                    left_label="7.1 matriculas de graduacao",
                    left_value=total_71,
                    right_label="5.1 total de graduacao",
                    right_value=total_51,
                    possible_cause=(
                        "Diferenca de cobertura, classificacao ou universo entre "
                        "as tabelas 7.1 e 5.1."
                    ),
                )
            )

            presential = direct_value(
                year, municipality_id, "enrollments_presential"
            )
            distance = direct_value(year, municipality_id, "enrollments_distance")
            modality_sum = (
                presential + distance
                if presential is not None and distance is not None
                else None
            )
            modality_check = _comparison(
                year=year,
                municipality_id=municipality_id,
                municipality=municipality,
                measure="enrollments_total_vs_modalities",
                left_label="7.1 matriculas totais",
                left_value=total_71,
                right_label="7.2 presencial + 7.3 EaD",
                right_value=modality_sum,
                possible_cause=(
                    "Cobertura territorial ou registro de modalidade nao "
                    "reconciliado entre 7.1, 7.2 e 7.3."
                ),
            )
            comparisons.append(modality_check)
            if modality_check["status"] != "matched":
                modality_unavailable.append(
                    {
                        "year": year,
                        "municipalityId": municipality_id,
                        "municipality": municipality,
                        "reason": modality_check["status"],
                    }
                )

            total_ies = direct_value(year, municipality_id, "ies_headquarters")
            total_faculty = direct_value(year, municipality_id, "faculty_total")
            checks = (
                (
                    "enrollment_dependencies",
                    total_51,
                    _sum_records(
                        records,
                        year=year,
                        municipality_id=municipality_id,
                        metric="enrollments_by_dependency",
                    ),
                    "5.1 total",
                    "5.1 dependencias",
                ),
                (
                    "enrollment_organizations",
                    total_51,
                    _sum_records(
                        records,
                        year=year,
                        municipality_id=municipality_id,
                        metric="enrollments_by_organization",
                    ),
                    "5.1 total",
                    "5.1 organizacoes",
                ),
                (
                    "ies_dependencies",
                    total_ies,
                    _sum_records(
                        records,
                        year=year,
                        municipality_id=municipality_id,
                        metric="ies_by_dependency",
                    ),
                    "1.1 total",
                    "1.1 dependencias",
                ),
                (
                    "ies_organizations",
                    total_ies,
                    _sum_records(
                        records,
                        year=year,
                        municipality_id=municipality_id,
                        metric="ies_by_organization",
                    ),
                    "1.1 total",
                    "1.1 organizacoes",
                ),
                (
                    "faculty_education",
                    total_faculty,
                    _sum_records(
                        records,
                        year=year,
                        municipality_id=municipality_id,
                        metric="faculty_by_education",
                    ),
                    "2.1 total",
                    "2.3 formacao",
                ),
            )
            for measure, left, right, left_label, right_label in checks:
                comparison = _comparison(
                    year=year,
                    municipality_id=municipality_id,
                    municipality=municipality,
                    measure=measure,
                    left_label=left_label,
                    left_value=left,
                    right_label=right_label,
                    right_value=right,
                    possible_cause=(
                        "Categorias ausentes, nao exaustivas ou classificadas "
                        "em universo distinto."
                    ),
                )
                comparisons.append(comparison)
                if measure == "faculty_education":
                    formation_exhaustiveness.append(
                        {
                            "year": year,
                            "municipalityId": municipality_id,
                            "municipality": municipality,
                            "exhaustive": comparison["status"] == "matched",
                        }
                    )

        state_71 = state_controls.get((year, "7.1"), {})
        state_51 = state_controls.get((year, "5.1"), {})
        state_72 = state_controls.get((year, "7.2"), {})
        state_73 = state_controls.get((year, "7.3"), {})
        state_21 = state_controls.get((year, "2.1"), {})
        state_23 = state_controls.get((year, "2.3"), {})
        state_checks = (
            (
                "state_enrollments_7.1_vs_5.1",
                state_71.get("enrollments_total"),
                state_51.get("enrollments_total_control"),
                f"7.1 {state_config.state_code}",
                f"5.1 {state_config.state_code}",
            ),
            (
                "state_enrollments_total_vs_modalities",
                state_71.get("enrollments_total"),
                (
                    state_72.get("enrollments_presential")
                    + state_73.get("enrollments_distance")
                    if state_72.get("enrollments_presential") is not None
                    and state_73.get("enrollments_distance") is not None
                    else None
                ),
                f"7.1 {state_config.state_code}",
                f"7.2 + 7.3 {state_config.state_code}",
            ),
            (
                "state_faculty_2.1_vs_2.3",
                state_21.get("faculty_total"),
                state_23.get("faculty_total_control"),
                f"2.1 {state_config.state_code}",
                f"2.3 {state_config.state_code}",
            ),
        )
        for measure, left, right, left_label, right_label in state_checks:
            comparisons.append(
                _comparison(
                    year=year,
                    municipality_id=state_config.state_code,
                    municipality=state_config.state_name,
                    measure=measure,
                    left_label=left_label,
                    left_value=left,
                    right_label=right_label,
                    right_value=right,
                    possible_cause=(
                        "Diferenca entre agregados oficiais de tabelas distintas."
                    ),
                    scope="state_control",
                )
            )

    return comparisons, formation_exhaustiveness, modality_unavailable


def _comparison_summary(comparisons: Sequence[dict]) -> dict:
    by_measure: dict[str, Counter] = defaultdict(Counter)
    for comparison in comparisons:
        by_measure[comparison["measure"]][comparison["status"]] += 1
    return {
        measure: dict(sorted(counter.items()))
        for measure, counter in sorted(by_measure.items())
    }


def _coverage(records: Sequence[NormalizedRecord]) -> dict:
    coverage: dict[str, dict[str, set[str]]] = defaultdict(
        lambda: defaultdict(set)
    )
    for record in records:
        if record.status in {STATUS_OBSERVED, STATUS_DERIVED_ZERO}:
            coverage[str(record.year)][record.metric].add(record.municipality_id)
    return {
        year: {
            metric: len(municipalities)
            for metric, municipalities in sorted(metrics.items())
        }
        for year, metrics in sorted(coverage.items())
    }


def _record_counts(records: Sequence[NormalizedRecord]) -> dict:
    counts = Counter((record.year, record.source_table) for record in records)
    return {
        str(year): {
            table: counts[(year, table)]
            for table in REQUIRED_TABLES
        }
        for year in sorted({record.year for record in records})
    }


def _layout_changes(layouts: Sequence[SheetLayout]) -> list[dict]:
    by_table: dict[str, list[SheetLayout]] = defaultdict(list)
    for layout in layouts:
        by_table[layout.table].append(layout)
    changes: list[dict] = []
    for table, table_layouts in sorted(by_table.items()):
        first = min(table_layouts, key=lambda item: item.year)
        latest = max(table_layouts, key=lambda item: item.year)
        if first.dimensions != latest.dimensions:
            changes.append(
                {
                    "table": table,
                    "firstYear": first.year,
                    "latestYear": latest.year,
                    "firstDimensions": first.dimensions,
                    "latestDimensions": latest.dimensions,
                    "comparabilityImpact": "none_semantic_locator",
                }
            )
    return changes


def _select_candidate(
    candidates: Iterable[str],
    used: set[str],
) -> str | None:
    ordered = sorted(set(candidates))
    for candidate in ordered:
        if candidate not in used:
            used.add(candidate)
            return candidate
    if ordered:
        return ordered[0]
    return None


def build_pilots(
    records: Sequence[NormalizedRecord],
    municipality_universe: Mapping[str, str],
    comparisons: Sequence[dict],
) -> list[dict]:
    latest_year = max(record.year for record in records)
    direct = _direct_index(records)
    latest_records = [record for record in records if record.year == latest_year]
    latest_any = {record.municipality_id for record in latest_records}

    def positive(metric: str) -> set[str]:
        return {
            record.municipality_id
            for record in latest_records
            if record.metric == metric
            and record.value is not None
            and record.value > 0
        }

    ies = positive("ies_headquarters")
    presential = positive("enrollments_presential")
    distance = positive("enrollments_distance")
    public_offer: set[str] = set()
    predominantly_private: set[str] = set()
    by_dependency: dict[str, dict[str, int]] = defaultdict(dict)
    for record in latest_records:
        if (
            record.metric == "enrollments_by_dependency"
            and record.value is not None
            and record.administrative_dependency
        ):
            by_dependency[record.municipality_id][
                record.administrative_dependency
            ] = record.value
    for municipality_id, values in by_dependency.items():
        public = sum(values.get(item, 0) for item in ("federal", "estadual", "municipal"))
        private = values.get("privada", 0)
        if public > 0:
            public_offer.add(municipality_id)
        if private > public:
            predominantly_private.add(municipality_id)

    used = {"4314902"}
    selections = [
        ("porto_alegre", "4314902", "Municipio fixo solicitado."),
        (
            "presential_offer",
            _select_candidate(presential, used),
            "Possui matriculas presenciais no ano mais recente.",
        ),
        (
            "distance_offer",
            _select_candidate(distance, used),
            "Possui matriculas EaD no ano mais recente.",
        ),
        (
            "distance_without_ies_headquarters",
            _select_candidate(distance - ies, used),
            "Possui oferta EaD e nao possui IES-sede no ano mais recente.",
        ),
        (
            "without_records",
            _select_candidate(set(municipality_universe) - latest_any, used),
            "Nao possui registros de Educacao Superior no ano mais recente.",
        ),
        (
            "public_offer",
            _select_candidate(public_offer, used),
            "Possui matriculas em dependencia publica no ano mais recente.",
        ),
        (
            "predominantly_private",
            _select_candidate(predominantly_private, used),
            "Matriculas privadas superam a soma federal, estadual e municipal.",
        ),
    ]

    pilots: list[dict] = []
    years = sorted({record.year for record in records})
    for role, municipality_id, reason in selections:
        if municipality_id is None:
            pilots.append(
                {
                    "role": role,
                    "municipalityId": None,
                    "selectionReason": reason,
                    "selectionStatus": "unavailable",
                }
            )
            continue
        municipality_records = [
            record
            for record in records
            if record.municipality_id == municipality_id
        ]
        indicators_found = sorted(
            {
                record.metric
                for record in municipality_records
                if record.metric in DIRECT_METRICS and record.value is not None
            }
        )
        years_by_indicator = {
            metric: sorted(
                {
                    record.year
                    for record in municipality_records
                    if record.metric == metric and record.value is not None
                }
            )
            for metric in indicators_found
        }
        universes = {
            record.metric: {
                "statisticalUniverse": record.statistical_universe,
                "territorialReference": record.territorial_reference,
            }
            for record in municipality_records
            if record.metric in DIRECT_METRICS
        }
        decompositions = sorted(
            {
                record.metric
                for record in municipality_records
                if record.metric in DECOMPOSITION_METRICS
                and record.value is not None
            }
        )
        zeroes = [
            {
                "year": record.year,
                "metric": record.metric,
                "status": record.status,
            }
            for record in municipality_records
            if record.value == 0
        ]
        absent = [
            {"year": year, "metric": metric}
            for year in years
            for metric in DIRECT_METRICS
            if (
                (year, municipality_id, metric) not in direct
                or direct[(year, municipality_id, metric)].value is None
            )
        ]
        municipality_comparisons = [
            comparison
            for comparison in comparisons
            if comparison["municipalityId"] == municipality_id
        ]
        pilots.append(
            {
                "role": role,
                "municipalityId": municipality_id,
                "municipality": municipality_universe[municipality_id],
                "selectionReason": reason,
                "selectionStatus": "selected",
                "indicatorsFound": indicators_found,
                "yearsByIndicator": years_by_indicator,
                "universes": universes,
                "decompositionsAvailable": decompositions,
                "observedAndDerivedZeroes": zeroes,
                "absences": absent,
                "reconciliations": municipality_comparisons,
                "divergences": [
                    item
                    for item in municipality_comparisons
                    if item["status"] == "mismatched"
                ],
            }
        )
    return pilots


def parse_higher_education_sources(
    *,
    source_dir: Path,
    state_config: StateConfig,
    registry: MunicipalityRegistry,
    years: Iterable[int] | None = None,
) -> ParsedAudit:
    source_paths = discover_source_files(source_dir, years)
    municipality_universe = municipality_universe_from_registry(registry)
    if registry.state_code != state_config.state_code:
        raise ValueError(
            "Registro municipal e configuração estadual possuem estados diferentes."
        )
    records: list[NormalizedRecord] = []
    layouts: list[SheetLayout] = []
    controls: dict[tuple[int, str, str], dict[str, int | None]] = {}
    state_controls: dict[tuple[int, str], dict[str, int | None]] = {}
    source_files: list[dict] = []
    sheet_coverage: dict[str, dict[str, int]] = {}

    for year, source_path in source_paths.items():
        digest = file_sha256(source_path)
        source_files.append(
            {
                "sourceId": f"inep-esup-{year}",
                "year": year,
                "fileName": source_path.name,
                "sha256": digest,
                "tables": list(REQUIRED_TABLES),
            }
        )
        workbook = load_workbook(source_path, read_only=True, data_only=True)
        try:
            names = _sheet_names(workbook)
            year_coverage: dict[str, int] = {}
            for table in REQUIRED_TABLES:
                parsed = _parse_sheet(
                    workbook,
                    year=year,
                    table=table,
                    sheet_name=names[table],
                    source_file=source_path,
                    source_sha256=digest,
                    state_config=state_config,
                    municipality_universe=municipality_universe,
                )
                records.extend(parsed.records)
                layouts.append(parsed.layout)
                year_coverage[table] = parsed.municipality_count
                for municipality_id, values in parsed.controls.items():
                    controls[(year, table, municipality_id)] = values
                state_controls[(year, table)] = parsed.state_controls
            sheet_coverage[str(year)] = year_coverage
        finally:
            workbook.close()

    derived_zeroes = apply_derived_zero_rule(records, source_files)
    _records_index(records)
    comparisons, formation_exhaustiveness, modality_unavailable = (
        build_reconciliations(
            records,
            controls,
            state_controls,
            municipality_universe,
            state_config,
        )
    )
    divergences = [
        comparison
        for comparison in comparisons
        if comparison["status"] == "mismatched"
    ]
    unavailable_comparisons = [
        comparison
        for comparison in comparisons
        if comparison["status"] == "unavailable"
    ]
    pilots = build_pilots(records, municipality_universe, comparisons)
    quality_report = {
        "schemaVersion": 1,
        "scope": "ESUP-1",
        "firstYear": min(source_paths),
        "latestYear": max(source_paths),
        "availableYears": sorted(source_paths),
        "requiredTables": list(REQUIRED_TABLES),
        "sourceFiles": source_files,
        "layouts": [asdict(layout) for layout in layouts],
        "layoutChanges": _layout_changes(layouts),
        "recordsByYearAndTable": _record_counts(records),
        "sourceRowsByYearAndTable": sheet_coverage,
        "municipalCoverageByIndicator": _coverage(records),
        "normalizedRecordCount": len(records),
        "derivedZeroes": derived_zeroes,
        "reconciliationSummary": _comparison_summary(comparisons),
        "reconciliations": comparisons,
        "divergences": divergences,
        "unavailableComparisons": unavailable_comparisons,
        "facultyEducationExhaustiveness": formation_exhaustiveness,
        "unavailableModalityBreakdowns": modality_unavailable,
        "comparabilityBreaks": [
            {
                "year": item["year"],
                "municipalityId": item["municipalityId"],
                "measure": "enrollment_modality_composition",
                "reason": item["reason"],
            }
            for item in modality_unavailable
        ],
        "stateControlsMaterializedPublicly": False,
    }
    records.sort(key=lambda record: record.key())
    return ParsedAudit(
        records=records,
        quality_report=quality_report,
        pilots=pilots,
        source_files=source_files,
    )


def write_audit_outputs(audit: ParsedAudit, output_dir: Path | None = None) -> Path:
    if output_dir is None:
        target = Path(tempfile.mkdtemp(prefix="esup-audit-"))
    else:
        target = output_dir.resolve()
        target.mkdir(parents=True, exist_ok=True)
    normalized_target = str(target).replace("\\", "/").lower()
    if "/public/data" in normalized_target:
        raise ValueError("A auditoria ESUP-1 nao pode escrever em public/data.")

    records_path = target / "normalized_records.jsonl"
    with records_path.open("w", encoding="utf-8", newline="\n") as destination:
        for record in audit.records:
            destination.write(
                json.dumps(asdict(record), ensure_ascii=False, sort_keys=True)
                + "\n"
            )
    (target / "quality_report.json").write_text(
        json.dumps(
            audit.quality_report,
            ensure_ascii=False,
            indent=2,
            sort_keys=True,
        )
        + "\n",
        encoding="utf-8",
    )
    (target / "pilots.json").write_text(
        json.dumps(audit.pilots, ensure_ascii=False, indent=2, sort_keys=True)
        + "\n",
        encoding="utf-8",
    )
    return target


def audit_summary(audit: ParsedAudit, output_dir: Path) -> dict:
    return {
        "scope": "ESUP-1",
        "outputDir": str(output_dir),
        "years": audit.quality_report["availableYears"],
        "tables": list(REQUIRED_TABLES),
        "normalizedRecords": len(audit.records),
        "derivedZeroes": len(audit.quality_report["derivedZeroes"]),
        "divergences": len(audit.quality_report["divergences"]),
        "unavailableComparisons": len(
            audit.quality_report["unavailableComparisons"]
        ),
        "pilots": [
            {
                "role": pilot["role"],
                "municipalityId": pilot.get("municipalityId"),
                "municipality": pilot.get("municipality"),
                "status": pilot["selectionStatus"],
            }
            for pilot in audit.pilots
        ],
    }
