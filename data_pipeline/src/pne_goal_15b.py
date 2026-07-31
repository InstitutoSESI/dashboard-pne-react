"""Relações da Meta 15.b a partir da tabela 2.2 da Sinopse Superior."""

from __future__ import annotations

import hashlib
import json
import math
import unicodedata
from pathlib import Path
from typing import Any, Mapping

from openpyxl import load_workbook


YEARS = tuple(range(2018, 2025))
EXPECTED_MUNICIPALITIES = 497
EXPECTED_ROWS_WITH_IES = {
    2018: 43,
    2019: 41,
    2020: 37,
    2021: 38,
    2022: 36,
    2023: 36,
    2024: 36,
}
RELATION_IDS = (
    "15.b.total",
    "15.b.universidades",
    "15.b.centros_universitarios",
    "15.b.faculdades",
)

# Pares 1-based (total, tempo integral) por categoria administrativa oficial.
ORGANIZATION_COLUMNS = {
    "15.b.universidades": (
        (12, 13),
        (18, 19),
        (24, 25),
        (31, 32),
        (37, 38),
    ),
    "15.b.centros_universitarios": (
        (44, 45),
        (50, 51),
        (56, 57),
        (63, 64),
        (69, 70),
    ),
    "15.b.faculdades": (
        (76, 77),
        (82, 83),
        (88, 89),
        (95, 96),
        (101, 102),
    ),
}

DATA_DIR = Path(__file__).resolve().parents[1] / "data"
SNAPSHOT_DIR = DATA_DIR / "pne_goal_15b"
UNIVERSE_PATH = DATA_DIR / "pne_goal_11b_census_2022" / "municipal_components.json"


def stable_json_bytes(value: Any) -> bytes:
    return (
        json.dumps(
            value,
            ensure_ascii=False,
            indent=2,
            sort_keys=True,
            allow_nan=False,
        )
        + "\n"
    ).encode("utf-8")


def sha256_bytes(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def _normal(value: object) -> str:
    decomposed = unicodedata.normalize("NFKD", str(value or ""))
    return " ".join(
        "".join(
            character
            for character in decomposed
            if not unicodedata.combining(character)
        )
        .casefold()
        .split()
    )


def _count(value: object, *, context: str) -> int:
    if value in (None, "", "-", "..", "..."):
        raise ValueError(f"Contagem ausente em {context}.")
    numeric = float(value)
    if not math.isfinite(numeric) or numeric < 0 or not numeric.is_integer():
        raise ValueError(f"Contagem inválida em {context}: {value!r}.")
    return int(numeric)


def load_universe(path: Path = UNIVERSE_PATH) -> dict[str, str]:
    rows = json.loads(path.read_text(encoding="utf-8"))
    output = {
        str(row["municipalityId"]): str(row["municipalityName"])
        for row in rows
    }
    if len(output) != EXPECTED_MUNICIPALITIES:
        raise ValueError("Universo da Meta 15.b deve conter 497 municípios.")
    return output


def _workbook_for_year(source_dir: Path, year: int) -> Path:
    matches = sorted(source_dir.glob(f"*{year}.xlsx"))
    if len(matches) != 1:
        raise FileNotFoundError(
            f"Esperado um workbook oficial único da Educação Superior {year}."
        )
    return matches[0]


def _semantic_headers(worksheet) -> None:
    title = _normal(worksheet.cell(4, 1).value)
    required = (
        "numero de docentes em exercicio",
        "organizacao academica",
        "dependencia administrativa",
        "regime de trabalho",
        "municipio",
    )
    if not all(fragment in title for fragment in required):
        raise ValueError("Título semântico inesperado na tabela 2.2.")
    checks = {
        (6, 1): "regiao geografica",
        (6, 2): "unidade da federacao",
        (6, 3): "municipio",
        (6, 4): "codigo do municipio",
        (7, 5): "total",
        (10, 6): "tempo integral",
        (7, 11): "universidade",
        (7, 43): "centro universitario",
        (7, 75): "faculdade",
    }
    for (row, column), expected in checks.items():
        if expected not in _normal(worksheet.cell(row, column).value):
            raise ValueError(
                f"Cabeçalho semântico divergente em L{row}C{column}."
            )


def parse_workbook(path: Path, *, year: int) -> dict[str, dict[str, list[int]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        matches = [name for name in workbook.sheetnames if name.strip() == "2.2"]
        if len(matches) != 1:
            raise ValueError(f"Tabela 2.2 não é única no workbook {year}.")
        worksheet = workbook[matches[0]]
        _semantic_headers(worksheet)
        output: dict[str, dict[str, list[int]]] = {}
        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=14, values_only=True),
            start=14,
        ):
            if _normal(row[1]) != "rio grande do sul":
                continue
            municipality_id = str(row[3] or "").strip().split(".")[0]
            municipality_name = str(row[2] or "").strip()
            if not municipality_id or not municipality_name:
                continue
            if municipality_id in output:
                raise ValueError(
                    f"Município duplicado na tabela 2.2/{year}: {municipality_id}."
                )
            indicators: dict[str, list[int]] = {}
            total_denominator = _count(
                row[4],
                context=f"{year}/linha {row_number}/total",
            )
            total_numerator = _count(
                row[5],
                context=f"{year}/linha {row_number}/tempo integral",
            )
            if total_numerator > total_denominator:
                raise ValueError(f"Numerador total excede denominador em {year}.")
            indicators["15.b.total"] = [total_numerator, total_denominator]
            for relation_id, pairs in ORGANIZATION_COLUMNS.items():
                denominator = sum(
                    _count(
                        row[total_column - 1],
                        context=f"{year}/linha {row_number}/C{total_column}",
                    )
                    for total_column, _ in pairs
                )
                numerator = sum(
                    _count(
                        row[full_time_column - 1],
                        context=f"{year}/linha {row_number}/C{full_time_column}",
                    )
                    for _, full_time_column in pairs
                )
                if numerator > denominator:
                    raise ValueError(
                        f"Numerador excede denominador em {relation_id}/{year}."
                    )
                indicators[relation_id] = [numerator, denominator]
            output[municipality_id] = {
                "municipalityName": municipality_name,
                "indicators": indicators,
            }
        expected = EXPECTED_ROWS_WITH_IES[year]
        if len(output) != expected:
            raise ValueError(
                f"Linhas municipais com IES em {year}: {len(output)} != {expected}."
            )
        return output
    finally:
        workbook.close()


def ratio(numerator: int, denominator: int) -> dict[str, Any]:
    if denominator == 0:
        return {
            "dataStatus": "not_applicable",
            "reasonCode": "denominator_zero",
            "value": None,
            "numerator": numerator,
            "denominator": denominator,
        }
    return {
        "dataStatus": "available",
        "value": 100.0 * numerator / denominator,
        "numerator": numerator,
        "denominator": denominator,
    }


def build_snapshot(
    source_dir: Path,
    *,
    reference_date: str,
    universe_path: Path = UNIVERSE_PATH,
) -> dict[str, bytes]:
    universe = load_universe(universe_path)
    parsed: dict[int, dict[str, dict[str, Any]]] = {}
    sources = []
    for year in YEARS:
        path = _workbook_for_year(source_dir.resolve(), year)
        parsed[year] = parse_workbook(path, year=year)
        sources.append(
            {
                "year": year,
                "table": "2.2",
                "fileName": path.name,
                "sha256": sha256_bytes(path.read_bytes()),
            }
        )

    municipal = []
    for municipality_id, municipality_name in sorted(universe.items()):
        series = []
        for year in YEARS:
            source = parsed[year].get(municipality_id)
            indicators = {}
            for relation_id in RELATION_IDS:
                numerator, denominator = (
                    source["indicators"][relation_id] if source else (0, 0)
                )
                indicators[relation_id] = ratio(numerator, denominator)
            series.append({"year": year, "indicators": indicators})
        municipal.append(
            {
                "municipalityId": municipality_id,
                "municipalityName": municipality_name,
                "series": series,
            }
        )

    state = []
    for year in YEARS:
        for relation_id in RELATION_IDS:
            numerator = sum(
                int(row["indicators"][relation_id][0])
                for row in parsed[year].values()
            )
            denominator = sum(
                int(row["indicators"][relation_id][1])
                for row in parsed[year].values()
            )
            state.append(
                {
                    "relationId": relation_id,
                    "territoryId": "43",
                    "territoryName": "Rio Grande do Sul",
                    "year": year,
                    **ratio(numerator, denominator),
                }
            )
    latest_total = next(
        row
        for row in state
        if row["year"] == 2024 and row["relationId"] == "15.b.total"
    )
    if (
        latest_total["numerator"] != 13755
        or latest_total["denominator"] != 22295
    ):
        raise ValueError("Reconciliação estadual total 15.b/2024 divergente.")

    municipal_bytes = stable_json_bytes(municipal)
    state_bytes = stable_json_bytes(state)
    manifest = {
        "schemaVersion": "pne-goal-15b-snapshot-v1",
        "sourceReferenceDate": reference_date,
        "years": list(YEARS),
        "municipalityCount": EXPECTED_MUNICIPALITIES,
        "table": "2.2",
        "unit": "docentes_em_exercicio_contabilizados_no_recorte",
        "territorialBasis": "municipality_of_institution_headquarters",
        "statePolicy": "ratio_of_municipal_sums",
        "organizationRecutPolicy": (
            "Soma das categorias administrativas oficiais dentro de cada "
            "organização; acompanhamento, não classificação legal por categoria."
        ),
        "sources": sources,
        "files": {
            "municipal_results.json": sha256_bytes(municipal_bytes),
            "state_results.json": sha256_bytes(state_bytes),
        },
    }
    return {
        "municipal_results.json": municipal_bytes,
        "state_results.json": state_bytes,
        "manifest.json": stable_json_bytes(manifest),
    }


def load_snapshot(
    snapshot_dir: Path = SNAPSHOT_DIR,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    root = snapshot_dir.resolve()
    manifest = json.loads((root / "manifest.json").read_text(encoding="utf-8"))
    if manifest.get("schemaVersion") != "pne-goal-15b-snapshot-v1":
        raise ValueError("Schema do snapshot da Meta 15.b inválido.")
    for filename, expected in (manifest.get("files") or {}).items():
        if sha256_bytes((root / filename).read_bytes()) != expected:
            raise ValueError(f"Hash divergente no snapshot 15.b: {filename}.")
    municipal = json.loads(
        (root / "municipal_results.json").read_text(encoding="utf-8")
    )
    state = json.loads((root / "state_results.json").read_text(encoding="utf-8"))
    if len(municipal) != EXPECTED_MUNICIPALITIES:
        raise ValueError("Cobertura municipal do snapshot 15.b inválida.")
    return municipal, state, manifest
