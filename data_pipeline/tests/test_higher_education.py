from __future__ import annotations

import json
import hashlib
import sys
import tempfile
import unittest
from dataclasses import asdict
from pathlib import Path
from unittest import mock

from openpyxl import Workbook


DATA_PIPELINE_DIR = Path(__file__).resolve().parents[1]
if str(DATA_PIPELINE_DIR) not in sys.path:
    sys.path.insert(0, str(DATA_PIPELINE_DIR))

from src.higher_education import (  # noqa: E402
    NormalizedRecord,
    ParsedAudit,
    STATUS_DERIVED_ZERO,
    STATUS_OBSERVED,
    apply_derived_zero_rule,
    discover_source_files,
    extract_year,
    file_sha256,
    normalize_ibge_code,
    parse_higher_education_sources,
    write_audit_outputs,
)
import src.higher_education_materialization as materialization  # noqa: E402
from src.higher_education_materialization import (  # noqa: E402
    BREAKDOWNS,
    INDICATORS,
    build_public_contracts,
    replace_directory_atomically,
    tree_hash,
    validate_public_directory,
    write_public_contracts,
)


PORTO_ALEGRE = "4314902"
TABLES = ("1.1", "2.1", "2.3", "5.1", "7.1", "7.2", "7.3")


def _path(*parts: str) -> tuple[str, ...]:
    return parts


def _sheet_spec(table: str, shifted: bool) -> tuple[str, list[tuple[str, tuple[str, ...]]]]:
    organization = "Organização Acadêmica e Dependência Administrativa"
    dimensions = [
        ("region", _path("Região Geográfica")),
        ("uf", _path("Unidade da Federação")),
    ]
    municipality = [
        ("municipality", _path("Município")),
        ("code", _path("Código do Município")),
    ]
    academic = [("level", _path("Nível Acadêmico do Curso"))]
    if table in {"5.1", "7.1"}:
        dimensions += municipality + academic if shifted else academic + municipality
    else:
        dimensions += municipality

    dependencies = [
        ("federal", _path(organization, "Total por Dependência Administrativa", "Pública", "Federal")),
        ("state", _path(organization, "Total por Dependência Administrativa", "Pública", "Estadual")),
        ("municipal", _path(organization, "Total por Dependência Administrativa", "Pública", "Municipal")),
        ("private", _path(organization, "Total por Dependência Administrativa", "Privada", "Total")),
    ]
    organizations = [
        ("university", _path(organization, "Universidade", "Total")),
        ("university_center", _path(organization, "Centro Universitário", "Total")),
        ("faculty", _path(organization, "Faculdade", "Total")),
        (
            "federal_institute",
            _path(organization, "Instituto Federal de Educação, Ciência e Tecnologia e Cefet", "Federal"),
        ),
    ]

    if table == "1.1":
        title = (
            "Número de Instituições de Ensino Superior por Município da Sede "
            "Administrativa - 2024"
        )
        metrics = [("total", _path(organization, "Total"))] + dependencies + organizations
    elif table == "2.1":
        title = "Número de Docentes em Exercício por Município - 2024"
        metrics = [("total", _path(organization, "Total"))]
    elif table == "2.3":
        title = (
            "Número de Docentes em Exercício por Município, Sexo e Grau de "
            "Formação - 2024"
        )
        formation = "Total por Sexo e por Grau de Formação"
        metrics = [
            (
                "total",
                _path(
                    "Organização Acadêmica, Dependência Administrativa, Sexo e Grau",
                    "Total",
                ),
            ),
            ("male", _path(formation, "Sexo", "Docentes", "Masculino")),
            ("female", _path(formation, "Sexo", "Docentes", "Feminino")),
            ("no_degree", _path(formation, "Grau de Formação", "Docentes", "Sem Graduação")),
            ("degree", _path(formation, "Grau de Formação", "Docentes", "Graduação")),
            ("specialization", _path(formation, "Grau de Formação", "Docentes", "Especialização")),
            ("masters", _path(formation, "Grau de Formação", "Docentes", "Mestrado")),
            ("doctorate", _path(formation, "Grau de Formação", "Docentes", "Doutorado")),
        ]
    elif table == "5.1":
        title = (
            "Número de Matrículas em Cursos de Graduação e Sequenciais de Formação "
            "Específica, Presenciais e a Distância - 2024"
        )
        metrics = [("total", _path(organization, "Total"))] + dependencies + organizations
    elif table == "7.1":
        title = "Número de Ingressantes, Matrículas e Concluintes - 2024"
        metrics = [
            ("entrants", _path("Total Geral", "Ingressantes")),
            ("enrollments", _path("Total Geral", "Matrículas")),
            ("graduates", _path("Total Geral", "Concluintes")),
        ]
    elif table == "7.2":
        title = "Cursos de Graduação Presenciais - 2024"
        metrics = [
            ("courses", _path("Total Geral", "Número de Cursos")),
            ("vacancies", _path("Total Geral", "Vagas Oferecidas")),
            ("enrollments", _path("Total Geral", "Matrículas")),
        ]
    else:
        title = "Cursos de Graduação a Distância e Número de Polos - 2024"
        metrics = [
            ("poles", _path("Total Geral", "Número de Polos")),
            ("courses", _path("Total Geral", "Número de Cursos")),
            ("enrollments", _path("Total Geral", "Matrículas")),
        ]
    return title, dimensions + metrics


def _values(table: str) -> dict[str, int]:
    common = {
        "total": 100,
        "federal": 10,
        "state": 20,
        "municipal": 30,
        "private": 40,
        "university": 30,
        "university_center": 20,
        "faculty": 40,
        "federal_institute": 10,
    }
    if table == "1.1":
        return {
            **common,
            "total": 2,
            "federal": 1,
            "state": 0,
            "municipal": 0,
            "private": 1,
            "university": 1,
            "university_center": 0,
            "faculty": 1,
            "federal_institute": 0,
        }
    if table == "2.1":
        return {"total": 10}
    if table == "2.3":
        return {
            "total": 10,
            "male": 6,
            "female": 4,
            "no_degree": 0,
            "degree": 1,
            "specialization": 2,
            "masters": 3,
            "doctorate": 4,
        }
    if table == "7.1":
        return {"entrants": 20, "enrollments": 100, "graduates": 10}
    if table == "7.2":
        return {"courses": 3, "vacancies": 50, "enrollments": 40}
    if table == "7.3":
        return {"poles": 2, "courses": 4, "enrollments": 60}
    return common


def _write_row(sheet, row_number: int, columns, values, *, uf, municipality, code, level="Graduação"):
    dimensions = {
        "region": "Sul",
        "uf": uf,
        "municipality": municipality,
        "code": code,
        "level": level,
    }
    for column, (key, _) in enumerate(columns, start=1):
        sheet.cell(row_number, column, dimensions.get(key, values.get(key)))


def _make_workbook(
    path: Path,
    *,
    shifted: bool = True,
    duplicate: bool = False,
    negative: bool = False,
    nonexhaustive_education: bool = False,
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for table in TABLES:
        title, columns = _sheet_spec(table, shifted)
        sheet = workbook.create_sheet(f"{table} teste")
        sheet.cell(1, 1, title)
        for column, (_, parts) in enumerate(columns, start=1):
            for offset, part in enumerate(parts):
                sheet.cell(6 + offset, column, part)
        values = _values(table)
        if negative and table == "7.1":
            values["enrollments"] = -1
        if nonexhaustive_education and table == "2.3":
            values["doctorate"] = 3
        _write_row(
            sheet,
            10,
            columns,
            values,
            uf="Total Brasil",
            municipality="Total Brasil",
            code=None,
        )
        _write_row(
            sheet,
            11,
            columns,
            values,
            uf="Total Rio Grande do Sul",
            municipality="Total Rio Grande do Sul",
            code=None,
        )
        _write_row(
            sheet,
            12,
            columns,
            values,
            uf="Rio Grande do Sul",
            municipality="Porto Alegre",
            code=int(PORTO_ALEGRE),
        )
        if table in {"5.1", "7.1"}:
            _write_row(
                sheet,
                13,
                columns,
                values,
                uf="Rio Grande do Sul",
                municipality="Porto Alegre",
                code=int(PORTO_ALEGRE),
                level="Pós-Graduação",
            )
        if duplicate and table == "7.1":
            _write_row(
                sheet,
                14,
                columns,
                values,
                uf="Rio Grande do Sul",
                municipality="Porto Alegre",
                code=int(PORTO_ALEGRE),
            )
        _write_row(
            sheet,
            15,
            columns,
            values,
            uf="Santa Catarina",
            municipality="Florianópolis",
            code=4205407,
        )
    workbook.save(path)


def _municipality_index(path: Path) -> None:
    municipalities = [
        {"id_municipio": PORTO_ALEGRE, "municipio": "Porto Alegre"}
    ]
    municipalities.extend(
        {
            "id_municipio": str(4300000 + index),
            "municipio": f"Município {index}",
        }
        for index in range(1, 497)
        if str(4300000 + index) != PORTO_ALEGRE
    )
    while len(municipalities) < 497:
        index = len(municipalities) + 500
        municipalities.append(
            {
                "id_municipio": str(4300000 + index),
                "municipio": f"Município {index}",
            }
        )
    path.write_text(
        json.dumps({"municipios": municipalities}, ensure_ascii=False),
        encoding="utf-8",
    )


def _record(metric: str, value: int | None, table: str) -> NormalizedRecord:
    modality = (
        "presential"
        if metric == "enrollments_presential"
        else "distance" if metric == "enrollments_distance" else None
    )
    return NormalizedRecord(
        year=2024,
        municipality_id=PORTO_ALEGRE,
        municipality="Porto Alegre",
        metric=metric,
        value=value,
        modality=modality,
        administrative_dependency=None,
        academic_organization=None,
        faculty_education=None,
        statistical_universe="graduation",
        territorial_reference="course_offer_location",
        status=STATUS_OBSERVED if value is not None else "unavailable",
        source_file="sinopse_2024.xlsx",
        source_table=table,
        semantic_header_path="Total Geral",
        source_sha256="a" * 64,
    )


def _source_files() -> list[dict]:
    return [
        {
            "sourceId": f"inep-esup-{year}",
            "year": year,
            "fileName": f"Sinopse_Educacao_Superior_Municipio_{year}.xlsx",
            "sha256": hashlib.sha256(str(year).encode("utf-8")).hexdigest(),
            "tables": list(TABLES),
        }
        for year in range(2018, 2025)
    ]


def _public_record(
    *,
    municipality_id: str,
    municipality: str,
    year: int,
    metric: str,
    value: int,
    status: str,
    source_table: str,
    dimension: str | None = None,
    category: str | None = None,
) -> NormalizedRecord:
    dimensions = {
        "administrative_dependency": None,
        "academic_organization": None,
        "faculty_education": None,
    }
    if dimension is not None:
        dimensions[dimension] = category
    spec = next(
        (
            item
            for item in (*INDICATORS, *BREAKDOWNS)
            if item.metric == metric
        ),
        None,
    )
    assert spec is not None
    modality = None
    if metric == "enrollments_presential":
        modality = "presential"
    elif metric == "enrollments_distance":
        modality = "distance"
    return NormalizedRecord(
        year=year,
        municipality_id=municipality_id,
        municipality=municipality,
        metric=metric,
        value=value,
        modality=modality,
        administrative_dependency=dimensions["administrative_dependency"],
        academic_organization=dimensions["academic_organization"],
        faculty_education=dimensions["faculty_education"],
        statistical_universe=spec.universe,
        territorial_reference=spec.territorial_reference,
        status=status,
        source_file=f"Sinopse_Educacao_Superior_Municipio_{year}.xlsx",
        source_table=source_table,
        semantic_header_path="Total",
        source_sha256=hashlib.sha256(str(year).encode("utf-8")).hexdigest(),
    )


def _materialization_audit() -> ParsedAudit:
    records: list[NormalizedRecord] = []
    for spec in INDICATORS:
        records.append(
            _public_record(
                municipality_id=PORTO_ALEGRE,
                municipality="Porto Alegre",
                year=2024,
                metric=spec.metric,
                value=0 if spec.metric == "enrollments_presential" else 10,
                status=(
                    STATUS_DERIVED_ZERO
                    if spec.metric == "enrollments_presential"
                    else STATUS_OBSERVED
                ),
                source_table=spec.source_table,
            )
        )
    records.append(
        _public_record(
            municipality_id="4300001",
            municipality="Município 1",
            year=2023,
            metric="enrollments_total",
            value=5,
            status=STATUS_OBSERVED,
            source_table="7.1",
        )
    )
    reconciliations = []
    for spec in BREAKDOWNS:
        for category_id, _label in spec.categories:
            records.append(
                _public_record(
                    municipality_id=PORTO_ALEGRE,
                    municipality="Porto Alegre",
                    year=2024,
                    metric=spec.metric,
                    value=1,
                    status=STATUS_OBSERVED,
                    source_table=spec.source_table,
                    dimension=spec.dimension,
                    category=category_id,
                )
            )
        reconciliations.append(
            {
                "scope": "municipality",
                "measure": spec.reconciliation_measure,
                "year": 2024,
                "municipalityId": PORTO_ALEGRE,
                "status": "matched",
            }
        )
    return ParsedAudit(
        records=records,
        quality_report={
            "availableYears": list(range(2018, 2025)),
            "reconciliations": reconciliations,
        },
        pilots=[],
        source_files=_source_files(),
    )


class HigherEducationParserTest(unittest.TestCase):
    def _parse(self, root: Path, **workbook_options):
        source_dir = root / "source"
        source_dir.mkdir(parents=True)
        workbook_path = source_dir / "sinopse_2024.xlsx"
        _make_workbook(workbook_path, **workbook_options)
        index_path = root / "municipios_index.json"
        _municipality_index(index_path)
        return parse_higher_education_sources(
            source_dir=source_dir,
            municipality_index_path=index_path,
            years=(2024,),
        )

    def test_discovers_year_and_hashes_source(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = root / "sinopse_2024.xlsx"
            source.write_bytes(b"xlsx")
            self.assertEqual(extract_year(source), 2024)
            self.assertEqual(discover_source_files(root, (2024,)), {2024: source})
            self.assertEqual(len(file_sha256(source)), 64)

    def test_discovery_rejects_missing_and_duplicate_years(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            with self.assertRaisesRegex(ValueError, "ausentes"):
                discover_source_files(root, (2024,))
            (root / "a_2024.xlsx").write_bytes(b"a")
            (root / "b_2024.xlsx").write_bytes(b"b")
            with self.assertRaisesRegex(ValueError, "Mais de um"):
                discover_source_files(root, (2024,))

    def test_semantic_headers_survive_position_change(self):
        signatures = []
        for shifted in (False, True):
            with tempfile.TemporaryDirectory() as directory:
                audit = self._parse(Path(directory), shifted=shifted)
                signatures.append(
                    [
                        (record.metric, record.value)
                        for record in audit.records
                        if record.municipality_id == PORTO_ALEGRE
                    ]
                )
        self.assertEqual(signatures[0], signatures[1])

    def test_filters_rs_municipalities_and_keeps_state_controls_internal(self):
        with tempfile.TemporaryDirectory() as directory:
            audit = self._parse(Path(directory))
        self.assertTrue(audit.records)
        self.assertEqual({record.municipality_id for record in audit.records}, {PORTO_ALEGRE})
        self.assertFalse(
            any(record.municipality_id == "RS" for record in audit.records)
        )
        self.assertTrue(
            any(
                item["scope"] == "state_control"
                for item in audit.quality_report["reconciliations"]
            )
        )

    def test_ibge_requires_seven_integer_digits(self):
        self.assertEqual(normalize_ibge_code(4314902.0, field="teste"), PORTO_ALEGRE)
        for invalid in (123456, 4314902.5, "não é código"):
            with self.assertRaises(ValueError):
                normalize_ibge_code(invalid, field="teste")

    def test_only_graduation_line_is_materialized(self):
        with tempfile.TemporaryDirectory() as directory:
            audit = self._parse(Path(directory))
        totals = [
            record
            for record in audit.records
            if record.metric == "enrollments_total"
        ]
        self.assertEqual(len(totals), 1)
        self.assertEqual(totals[0].value, 100)

    def test_faculty_education_uses_total_block_without_sex_double_count(self):
        with tempfile.TemporaryDirectory() as directory:
            audit = self._parse(Path(directory))
        education = [
            record
            for record in audit.records
            if record.metric == "faculty_by_education"
        ]
        self.assertEqual(len(education), 5)
        self.assertEqual(sum(record.value or 0 for record in education), 10)
        self.assertIn("Sem Graduação", {record.faculty_education for record in education})

    def test_derived_zero_requires_observed_total_and_equal_other_modality(self):
        records = [
            _record("enrollments_total", 60, "7.1"),
            _record("enrollments_distance", 60, "7.3"),
        ]
        derived = apply_derived_zero_rule(
            records,
            [{"year": 2024, "fileName": "sinopse_2024.xlsx", "sha256": "b" * 64}],
        )
        self.assertEqual(len(derived), 1)
        presential = next(
            record for record in records if record.metric == "enrollments_presential"
        )
        self.assertEqual(presential.value, 0)
        self.assertEqual(presential.status, STATUS_DERIVED_ZERO)

    def test_missing_or_divergent_modality_is_not_derived(self):
        cases = (
            [_record("enrollments_total", None, "7.1"), _record("enrollments_distance", 0, "7.3")],
            [_record("enrollments_total", 61, "7.1"), _record("enrollments_distance", 60, "7.3")],
        )
        for records in cases:
            derived = apply_derived_zero_rule(
                records,
                [{"year": 2024, "fileName": "sinopse_2024.xlsx", "sha256": "b" * 64}],
            )
            self.assertEqual(derived, [])

    def test_duplicate_municipality_is_rejected(self):
        with tempfile.TemporaryDirectory() as directory:
            with self.assertRaisesRegex(ValueError, "duplicado"):
                self._parse(Path(directory), duplicate=True)

    def test_negative_value_is_rejected(self):
        with tempfile.TemporaryDirectory() as directory:
            with self.assertRaisesRegex(ValueError, "negativo"):
                self._parse(Path(directory), negative=True)

    def test_output_is_deterministic(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            first = self._parse(root)
            second = parse_higher_education_sources(
                source_dir=root / "source",
                municipality_index_path=root / "municipios_index.json",
                years=(2024,),
            )
        self.assertEqual(
            [asdict(record) for record in first.records],
            [asdict(record) for record in second.records],
        )
        self.assertEqual(first.quality_report, second.quality_report)

    def test_reconciliations_match_complete_synthetic_source(self):
        with tempfile.TemporaryDirectory() as directory:
            audit = self._parse(Path(directory))
        municipal = [
            item
            for item in audit.quality_report["reconciliations"]
            if item["scope"] == "municipality"
        ]
        self.assertEqual({item["status"] for item in municipal}, {"matched"})
        self.assertEqual(audit.quality_report["divergences"], [])

    def test_nonexhaustive_faculty_education_is_flagged_without_adjustment(self):
        with tempfile.TemporaryDirectory() as directory:
            audit = self._parse(Path(directory), nonexhaustive_education=True)
        comparison = next(
            item
            for item in audit.quality_report["reconciliations"]
            if item["measure"] == "faculty_education"
            and item["scope"] == "municipality"
        )
        self.assertEqual(comparison["status"], "mismatched")
        self.assertFalse(audit.quality_report["facultyEducationExhaustiveness"][0]["exhaustive"])

    def test_audit_writer_rejects_public_data(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            audit = self._parse(root / "input")
            with self.assertRaisesRegex(ValueError, "public/data"):
                write_audit_outputs(audit, root / "public" / "data" / "educacao")


class HigherEducationMaterializationTest(unittest.TestCase):
    def _contracts(self, root: Path):
        index_path = root / "municipios_index.json"
        _municipality_index(index_path)
        universe = materialization.load_municipality_universe(index_path)
        manifest, municipalities = build_public_contracts(
            _materialization_audit(),
            universe,
        )
        return universe, manifest, municipalities

    def test_generates_manifest_and_exactly_497_municipal_files(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            universe, manifest, municipalities = self._contracts(root)
            output = root / "superior"
            write_public_contracts(output, manifest, municipalities)
            validation = validate_public_directory(output, universe)

            self.assertEqual(manifest["schemaVersion"], 1)
            self.assertEqual(manifest["municipalityCount"], 497)
            self.assertEqual(len(manifest["indicators"]), 9)
            self.assertEqual(len(manifest["breakdowns"]), 5)
            self.assertEqual(len(manifest["sources"]), 49)
            self.assertEqual(validation["municipalityFileCount"], 497)
            self.assertLess(validation["sizeBytes"]["manifest"], 100 * 1024)
            self.assertLess(
                validation["sizeBytes"]["municipalMaximum"],
                150 * 1024,
            )

    def test_availability_distinguishes_current_historical_and_unavailable(self):
        with tempfile.TemporaryDirectory() as directory:
            _universe, _manifest, municipalities = self._contracts(Path(directory))

        self.assertEqual(
            municipalities[PORTO_ALEGRE]["availability"],
            "current",
        )
        self.assertEqual(
            municipalities["4300001"]["availability"],
            "historical_only",
        )
        self.assertEqual(
            municipalities["4300002"]["availability"],
            "unavailable",
        )

    def test_series_preserve_gaps_derived_zero_and_source_ids(self):
        with tempfile.TemporaryDirectory() as directory:
            _universe, _manifest, municipalities = self._contracts(Path(directory))

        presential = municipalities[PORTO_ALEGRE]["indicators"][
            "esup-matriculas-presenciais"
        ]["series"]
        self.assertEqual([point["year"] for point in presential], list(range(2018, 2025)))
        self.assertTrue(
            all(
                point["status"] == "unavailable"
                and point["value"] is None
                and point["sourceId"] is None
                for point in presential[:-1]
            )
        )
        self.assertEqual(presential[-1]["status"], "derived_zero")
        self.assertEqual(presential[-1]["value"], 0)
        self.assertEqual(
            presential[-1]["sourceIds"],
            ["s-7.1-2024", "s-7.3-2024"],
        )

    def test_breakdowns_are_canonical_ordered_and_exhaustive_only_when_reconciled(self):
        with tempfile.TemporaryDirectory() as directory:
            _universe, _manifest, municipalities = self._contracts(Path(directory))

        breakdowns = municipalities[PORTO_ALEGRE]["breakdowns"]
        self.assertEqual(
            [(item["id"], item["year"]) for item in breakdowns],
            sorted(
                (
                    (spec.id, year)
                    for spec in BREAKDOWNS
                    for year in range(2018, 2025)
                )
            ),
        )
        current = [
            item
            for item in breakdowns
            if item["year"] == 2024
        ]
        self.assertTrue(all(item["exhaustive"] for item in current))
        self.assertTrue(
            all(
                not item["exhaustive"]
                for item in breakdowns
                if item["year"] < 2024
            )
        )

    def test_contract_has_no_local_paths_repeated_hashes_or_excluded_fields(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            universe, manifest, municipalities = self._contracts(root)
            output = root / "superior"
            write_public_contracts(output, manifest, municipalities)
            validate_public_directory(output, universe)
            municipal_text = (output / "municipios" / f"{PORTO_ALEGRE}.json").read_text(
                encoding="utf-8"
            )

        self.assertNotIn("sha256", municipal_text)
        self.assertNotIn("fileName", municipal_text)
        self.assertNotIn("latestValue", municipal_text)
        self.assertNotIn(str(root), municipal_text)

    def test_coverage_is_individual_and_source_registry_is_central(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            universe, manifest, municipalities = self._contracts(root)
            output = root / "superior"
            write_public_contracts(output, manifest, municipalities)
            validation = validate_public_directory(output, universe)

        coverage = {
            item["id"]: item["coverageByYear"]
            for item in manifest["indicators"]
        }
        self.assertEqual(
            coverage["esup-matriculas-total"]["2024"],
            1,
        )
        self.assertEqual(
            coverage["esup-vagas-presenciais"]["2023"],
            0,
        )
        self.assertEqual(validation["sourceIds"]["invalid"], 0)
        self.assertEqual(validation["indicatorFiles"], {
            spec.id: 497 for spec in INDICATORS
        })
        self.assertEqual(validation["breakdownFiles"], {
            spec.id: 497 for spec in BREAKDOWNS
        })

    def test_two_generations_are_byte_deterministic(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            universe, first_manifest, first_municipalities = self._contracts(root)
            _universe, second_manifest, second_municipalities = self._contracts(root)
            first = root / "first"
            second = root / "second"
            write_public_contracts(first, first_manifest, first_municipalities)
            write_public_contracts(second, second_manifest, second_municipalities)
            validate_public_directory(first, universe)
            validate_public_directory(second, universe)

            self.assertEqual(first_manifest["dataVersion"], second_manifest["dataVersion"])
            self.assertEqual(tree_hash(first), tree_hash(second))

    def test_atomic_replace_promotes_complete_stage(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            output = root / "superior"
            output.mkdir()
            (output / "old.json").write_text("old", encoding="utf-8")
            stage = root / "stage"
            stage.mkdir()
            (stage / "new.json").write_text("new", encoding="utf-8")

            replace_directory_atomically(stage, output)

            self.assertFalse((output / "old.json").exists())
            self.assertEqual(
                (output / "new.json").read_text(encoding="utf-8"),
                "new",
            )

    def test_atomic_failure_restores_previous_output(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            output = root / "superior"
            output.mkdir()
            (output / "old.json").write_text("old", encoding="utf-8")
            stage = root / "stage"
            stage.mkdir()
            (stage / "new.json").write_text("new", encoding="utf-8")
            real_rename = materialization.os.rename
            call_count = 0

            def fail_promotion(source, target):
                nonlocal call_count
                call_count += 1
                if call_count == 2:
                    raise OSError("falha simulada")
                return real_rename(source, target)

            with mock.patch.object(
                materialization.os,
                "rename",
                side_effect=fail_promotion,
            ):
                with self.assertRaisesRegex(OSError, "falha simulada"):
                    replace_directory_atomically(
                        stage,
                        output,
                        rename_attempts=1,
                    )

            self.assertEqual(
                (output / "old.json").read_text(encoding="utf-8"),
                "old",
            )
            self.assertFalse((output / "new.json").exists())

    def test_atomic_replace_retries_transient_windows_lock(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            output = root / "superior"
            stage = root / "stage"
            stage.mkdir()
            (stage / "new.json").write_text("new", encoding="utf-8")
            real_rename = materialization.os.rename
            call_count = 0

            def transient_lock(source, target):
                nonlocal call_count
                call_count += 1
                if call_count < 3:
                    raise PermissionError("bloqueio transitório")
                return real_rename(source, target)

            with mock.patch.object(
                materialization.os,
                "rename",
                side_effect=transient_lock,
            ), mock.patch.object(materialization.time, "sleep"):
                replace_directory_atomically(stage, output, rename_attempts=3)

            self.assertEqual(
                (output / "new.json").read_text(encoding="utf-8"),
                "new",
            )


if __name__ == "__main__":
    unittest.main()
