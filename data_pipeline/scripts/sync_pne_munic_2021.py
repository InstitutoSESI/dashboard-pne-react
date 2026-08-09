#!/usr/bin/env python3
"""Incorpora a MUNIC 2021 e audita o módulo climático da MUNIC 2024."""

from __future__ import annotations

import argparse
from collections import Counter
from hashlib import sha256
from pathlib import Path
import sys
import tempfile

import openpyxl
import requests


DATA_PIPELINE_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(DATA_PIPELINE_DIR))

from src.pne_macro_ingestion import (  # noqa: E402
    DATA_ROOT,
    MANIFEST_SCHEMA,
    canonical_json_bytes,
    load_municipality_universe,
    normalized_snapshot,
    normalize_name,
    raw_file_entry,
    write_source_snapshot,
)
from src.pne_state_context import (  # noqa: E402
    load_pne_state_context,
    resolve_state_snapshot_dir,
)


SOURCE_ID = "ibge_munic_2021"
URL_2021 = (
    "https://ftp.ibge.gov.br/Perfil_Municipios/2021/Base_de_Dados/"
    "Base_MUNIC_2021_20240425.xlsx"
)
URL_2024 = (
    "https://ftp.ibge.gov.br/Perfil_Municipios/2024/Base_de_Dados/"
    "Base_MUNIC_2024_20251107.xlsx"
)
SOURCE_PAGE = "https://www.ibge.gov.br/estatisticas/sociais/educacao/10586-pesquisa-de-informacoes-basicas-municipais.html"
DESTINATION = DATA_ROOT / "munic_2021"

RESPONSE_MAP = {
    "sim": "yes",
    "nao": "no",
    "-": "unknown",
    "": "unknown",
    "recusa": "unknown",
    "nao sabe informar": "unknown",
    "nao soube informar": "unknown",
    "nao ha professores com jornada de 40 horas semanais": "not_applicable",
}


def _download(url: str, destination: Path) -> Path:
    response = requests.get(url, timeout=180)
    response.raise_for_status()
    destination.write_bytes(response.content)
    return destination


def _response(value: object) -> str:
    key = normalize_name(value)
    if str(value or "").strip() == "-":
        key = "-"
    if key not in RESPONSE_MAP:
        raise ValueError(f"Resposta MUNIC não reconhecida: {value!r}")
    return RESPONSE_MAP[key]


def _dictionary_rows(workbook: openpyxl.Workbook) -> list[list[object]]:
    worksheet = workbook.worksheets[0]
    return [list(row) for row in worksheet.iter_rows(values_only=True)]


def parse_munic(
    source_2021: Path,
    source_2024: Path,
    state_code: str = "RS",
) -> tuple[dict, dict]:
    state = load_pne_state_context(state_code)
    municipality_names, _ = load_municipality_universe(state.state_code)
    workbook = openpyxl.load_workbook(
        source_2021,
        read_only=True,
        data_only=True,
    )
    try:
        education = next(
            sheet
            for sheet in workbook.worksheets
            if normalize_name(sheet.title) == "educacao"
        )
        rows = education.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(rows)]
        positions = {header.casefold(): index for index, header in enumerate(headers)}
        required = {
            "codmun",
            "uf",
            "mun",
            "medu15",
            "medu16",
            "medu18",
            "medu20a",
            "medu21",
        }
        missing = sorted(required - set(positions))
        if missing:
            raise ValueError(f"MUNIC 2021 sem variáveis obrigatórias: {missing}")

        records: dict[str, dict] = {}
        response_counts: dict[str, Counter[str]] = {
            field: Counter()
            for field in ("Medu15", "Medu16", "Medu18", "Medu20a", "Medu21")
        }
        for row in rows:
            if str(row[positions["uf"]] or "").strip() != state.state_code:
                continue
            municipality_id = str(row[positions["codmun"]] or "").split(".")[0]
            if municipality_id not in municipality_names:
                raise ValueError(
                    f"Código MUNIC fora do universo {state.state_code}: "
                    f"{municipality_id}"
                )
            if municipality_id in records:
                raise ValueError(f"Município duplicado na MUNIC 2021: {municipality_id}")
            values = {
                field: _response(row[positions[field.casefold()]])
                for field in response_counts
            }
            for field, value in values.items():
                response_counts[field][value] += 1
            records[municipality_id] = {
                "municipalityId": municipality_id,
                "municipalityName": municipality_names[municipality_id],
                "year": 2021,
                "careerPlans": {
                    "teacherPlan": values["Medu16"],
                    "twoThirdsInteractionLimit": values["Medu18"],
                    "initialSalaryAtLeast2021Floor": values["Medu20a"],
                    "nonTeachingPlan": values["Medu21"],
                },
                "educationForum": values["Medu15"],
            }
        dictionary_rows = _dictionary_rows(workbook)
    finally:
        workbook.close()

    normalized = normalized_snapshot(
        source_id=SOURCE_ID,
        edition="MUNIC 2021 — Educação",
        records=records,
        municipality_names=municipality_names,
        state_code=state.state_code,
    )

    climate_workbook = openpyxl.load_workbook(
        source_2024,
        read_only=True,
        data_only=True,
    )
    try:
        climate_dictionary = "\n".join(
            " | ".join(str(value or "") for value in row)
            for row in _dictionary_rows(climate_workbook)
        )
        climate_key = normalize_name(climate_dictionary)
        climate_plan_matches = [
            phrase
            for phrase in (
                "plano climatico",
                "plano de mudanca do clima",
                "plano de prevencao mitigacao e adaptacao",
            )
            if phrase in climate_key
        ]
    finally:
        climate_workbook.close()

    dictionary_variables = {
        "MEDU15": "Instituiu/constituiu Fórum Permanente de Educação",
        "MEDU16": "Plano de Carreira para o Magistério — existência",
        "MEDU18": (
            "Lei do Plano de Carreira prevê expressamente o limite de 2/3 "
            "da carga horária para interação com os educandos"
        ),
        "MEDU20a": (
            "Professores com jornada de 40 horas possuem vencimento básico "
            "inicial igual ou superior ao piso informado na edição"
        ),
        "MEDU21": (
            "Plano de Carreira vigente para profissionais da educação "
            "não docentes"
        ),
    }
    coverage = {
        "municipalityCount": len(records),
        "responseCounts": {
            field: dict(sorted(counts.items()))
            for field, counts in sorted(response_counts.items())
        },
    }
    audit = {
        "dictionaryRowCount": len(dictionary_rows),
        "dictionaryVariables": dictionary_variables,
        "coverage": coverage,
        "climatePlan2024": {
            "status": "blocked",
            "matches": climate_plan_matches,
            "reason": (
                "A MUNIC 2024 não contém variável sobre plano municipal ou "
                "da rede de ensino para prevenção, mitigação e adaptação "
                "às mudanças do clima."
            ),
        },
    }
    return normalized, audit


def materialize(
    source_2021: Path,
    source_2024: Path,
    state_code: str = "RS",
) -> dict:
    state = load_pne_state_context(state_code)
    normalized, audit = parse_munic(
        source_2021, source_2024, state.state_code
    )
    normalized_hash = sha256(canonical_json_bytes(normalized)).hexdigest()
    manifest = {
        "schemaVersion": MANIFEST_SCHEMA,
        "stateCode": state.state_code,
        "stateName": state.state_name,
        "sourceId": SOURCE_ID,
        "sourceTitle": "Pesquisa de Informações Básicas Municipais — MUNIC",
        "organization": "Instituto Brasileiro de Geografia e Estatística (IBGE)",
        "edition": "2021 (Educação); 2024 (auditoria climática)",
        "sourcePageUrl": SOURCE_PAGE,
        "rawFiles": [
            raw_file_entry(
                logical_name="MUNIC 2021",
                path=source_2021,
                official_url=URL_2021,
            ),
            raw_file_entry(
                logical_name="MUNIC 2024",
                path=source_2024,
                official_url=URL_2024,
            ),
        ],
        "dictionary": audit["dictionaryVariables"],
        "coverage": audit["coverage"],
        "absencePolicy": {
            "recusa": "unavailable",
            "nao_sabe_informar": "unavailable",
            "empty_or_dash": "unavailable",
            "explicit_no": "observed_zero_or_negative_declaration",
        },
        "duplicatePolicy": "código municipal duplicado invalida a carga",
        "normalization": {
            "municipalityKey": "CodMun (código IBGE de sete dígitos)",
            "responseMapping": RESPONSE_MAP,
            "normalizedSha256": normalized_hash,
        },
        "blocked": {"climatePlan2024": audit["climatePlan2024"]},
        "status": "partially_approved",
        "approvedComponents": ["careerPlans2021", "educationForum2021"],
    }
    write_source_snapshot(
        destination=resolve_state_snapshot_dir(DESTINATION, state.state_code),
        raw_files={
            "Base_MUNIC_2021_20240425.xlsx": source_2021,
            "Base_MUNIC_2024_20251107.xlsx": source_2024,
        },
        normalized=normalized,
        manifest=manifest,
    )
    return manifest


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--state", default="RS")
    parser.add_argument("--source-2021", type=Path)
    parser.add_argument("--source-2024", type=Path)
    args = parser.parse_args()
    with tempfile.TemporaryDirectory(
        prefix="pne-munic-",
        ignore_cleanup_errors=True,
    ) as temporary:
        root = Path(temporary)
        source_2021 = args.source_2021 or _download(
            URL_2021, root / "Base_MUNIC_2021_20240425.xlsx"
        )
        source_2024 = args.source_2024 or _download(
            URL_2024, root / "Base_MUNIC_2024_20251107.xlsx"
        )
        manifest = materialize(source_2021, source_2024, args.state)
    print(canonical_json_bytes(manifest).decode("utf-8"), end="")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
