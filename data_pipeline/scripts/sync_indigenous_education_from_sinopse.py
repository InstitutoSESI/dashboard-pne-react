"""Valida e importa o retrato municipal da Educação Escolar Indígena.

O importador lê diretamente as tabelas temáticas da Sinopse Estatística do
Censo Escolar, usa o código IBGE de sete dígitos como chave e materializa uma
tabela longa por município, ano, unidade de medida e recorte. Por padrão, a
execução apenas valida e imprime um resumo auditável. Use ``--apply`` para
substituir, em uma única transação, os anos selecionados na tabela
``educacao_indigena_municipal``.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import sys
import unicodedata
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from sqlalchemy import text

DATA_PIPELINE_DIR = Path(__file__).resolve().parents[1]
if str(DATA_PIPELINE_DIR) not in sys.path:
    sys.path.insert(0, str(DATA_PIPELINE_DIR))

from src.config import SESI_DB_DIR  # noqa: E402
from src.state_config import load_state_config  # noqa: E402

sys.path.insert(0, str(SESI_DB_DIR))
from utils_educacao import get_engine  # noqa: E402


RS_MUNICIPALITIES = 497
SUPPORTED_YEARS = (2023, 2024, 2025)
TARGET_UF = "Rio Grande do Sul"
TARGET_STATE_CODE = "RS"
TABLE_NAME = "educacao_indigena_municipal"

CUTS = (
    ("total", "Total"),
    ("educacao_infantil", "Educação Infantil"),
    ("creche", "Creche"),
    ("pre_escola", "Pré-Escola"),
    ("ensino_fundamental", "Ensino Fundamental"),
    ("anos_iniciais", "Anos Iniciais"),
    ("anos_finais", "Anos Finais"),
    ("ensino_medio", "Ensino Médio"),
    ("educacao_profissional", "Educação Profissional"),
    ("eja", "Educação de Jovens e Adultos"),
    ("eja_ensino_fundamental", "EJA - Ensino Fundamental"),
    ("eja_ensino_medio", "EJA - Ensino Médio"),
    ("educacao_especial", "Educação Especial"),
    ("classes_comuns", "Classes Comuns"),
    ("classes_exclusivas", "Classes Exclusivas"),
)

UNITS = {
    "matriculas": {
        "label": "Matrículas",
        "old_sheet": "1.52",
        "new_sheet": "1.74",
        "old_columns": (5, 6, 7, 8, 9, 10, 11, 12, 16, 24, 25, 26, 27, 28, 29),
        "new_columns": (5, 6, 7, 8, 9, 10, 11, 12, 17, 30, 31, 32, 33, 34, 35),
    },
    "docentes": {
        "label": "Docentes",
        "old_sheet": "2.56",
        "new_sheet": "2.77",
        "old_columns": (5, 6, 7, 8, 9, 10, 11, 12, 16, 25, 26, 27, 28, 29, 30),
        "new_columns": (5, 6, 7, 8, 9, 10, 11, 12, 18, 32, 33, 34, 35, 36, 37),
    },
    "estabelecimentos": {
        "label": "Estabelecimentos",
        "old_sheet": "3.34",
        "new_sheet": "3.56",
        "old_columns": (5, 6, 7, 8, 9, 10, 11, 12, 16, 24, 25, 26, 27, 28, 29),
        "new_columns": (5, 6, 7, 8, 9, 10, 11, 12, 17, 30, 31, 32, 33, 34, 35),
    },
    "turmas": {
        "label": "Turmas",
        "old_sheet": "4.23",
        "new_sheet": "4.52",
        "old_columns": (5, 6, 7, 8, 9, 10, 11, 12, 16, 25, 26, 27, 28, 29, 30),
        "new_columns": (5, 6, 7, 8, 9, 10, 11, 12, 18, 32, 33, 34, 35, 36, 37),
    },
}

HEADER_TOKENS = {
    "total": ("total",),
    "educacao_infantil": ("educacao infantil", "total"),
    "creche": ("creche",),
    "pre_escola": ("pre escola",),
    "ensino_fundamental": ("ensino fundamental", "total"),
    "anos_iniciais": ("anos iniciais",),
    "anos_finais": ("anos finais",),
    "ensino_medio": ("ensino medio", "total"),
    "educacao_profissional": ("educacao profissional", "total"),
    "eja": ("educacao de jovens e adultos", "total"),
    "eja_ensino_fundamental": ("ensino fundamental",),
    "eja_ensino_medio": ("ensino medio",),
    "educacao_especial": ("educacao especial", "total"),
    "classes_comuns": ("classes comuns",),
    "classes_exclusivas": ("classes exclusivas",),
}

COMPARABILITY_GROUPS = {
    "total": "comparavel_2023_2025",
    "educacao_infantil": "comparavel_2023_2025",
    "ensino_fundamental": "comparavel_2023_2025",
    "eja": "comparavel_2023_2025",
    "ensino_medio": "quebra_conceitual_2025",
    "educacao_profissional": "quebra_conceitual_2025",
    "educacao_especial": "quebra_conceitual_2025",
}

REFERENCE_VALUES_2025 = {
    "4315404": {"matriculas": 1029, "docentes": 78, "estabelecimentos": 9, "turmas": 82},
    "4304713": {"matriculas": 3, "docentes": 1, "estabelecimentos": 1, "turmas": 1},
    "4300034": {"matriculas": 0, "docentes": 0, "estabelecimentos": 0, "turmas": 0},
}


def _normalise(value: object) -> str:
    return " ".join(str(value or "").strip().split())


def _normalise_header(value: object) -> str:
    text_value = unicodedata.normalize("NFKD", _normalise(value))
    text_value = "".join(char for char in text_value if not unicodedata.combining(char))
    return " ".join(
        "".join(char if char.isalnum() else " " for char in text_value.lower()).split()
    )


def _optional_integer(value: object, *, field: str, row_number: int) -> int | None:
    if value is None or str(value).strip() in {"", "-", "—", "..."}:
        return None
    if isinstance(value, bool):
        raise ValueError(f"Valor inesperado em {field}, linha {row_number}: {value!r}.")
    try:
        numeric_float = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(
            f"Valor inválido em {field}, linha {row_number}: {value!r}."
        ) from exc
    if not math.isfinite(numeric_float) or not numeric_float.is_integer():
        raise ValueError(
            f"Valor não inteiro em {field}, linha {row_number}: {value!r}."
        )
    numeric = int(numeric_float)
    if numeric < 0:
        raise ValueError(f"Valor negativo em {field}, linha {row_number}.")
    return numeric


def _ibge_code(value: object, *, row_number: int) -> str:
    numeric = _optional_integer(value, field="codigo_ibge", row_number=row_number)
    if numeric is None or numeric < 1_000_000 or numeric > 9_999_999:
        raise ValueError(
            f"Código IBGE inválido na linha {row_number}: {value!r}."
        )
    return str(numeric)


def _file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _source_path(source_dir: Path, year: int) -> Path:
    candidates = sorted(
        path
        for path in source_dir.glob(f"*{year}*.xlsx")
        if not path.name.startswith("~$")
    )
    if len(candidates) != 1:
        raise ValueError(
            f"Esperada uma planilha .xlsx para {year} em {source_dir}; "
            f"encontradas {len(candidates)}."
        )
    return candidates[0]


def _sheet_name(workbook, suffix: str) -> str:
    matches = [name for name in workbook.sheetnames if name.endswith(suffix)]
    if len(matches) != 1:
        raise ValueError(
            f"Esperada uma aba terminada em {suffix}; encontradas {matches!r}."
        )
    return matches[0]


def _column_header(header_rows: list[tuple], column: int) -> str:
    index = column - 1
    return " ".join(
        _normalise_header(row[index])
        for row in header_rows
        if index < len(row) and _normalise_header(row[index])
    )


def _validate_layout_headers(
    header_rows: list[tuple],
    *,
    unit_key: str,
    columns: tuple[int, ...],
    sheet_name: str,
) -> None:
    blob = " ".join(_normalise_header(value) for row in header_rows for value in row)
    for required in ("educacao indigena", "codigo do municipio", "etapa de ensino"):
        if required not in blob:
            raise ValueError(
                f"Cabeçalho sem o marcador semântico {required!r} na aba {sheet_name}."
            )
    unit_label = _normalise_header(UNITS[unit_key]["label"])
    if unit_label not in blob:
        raise ValueError(
            f"Cabeçalho da aba {sheet_name} não identifica a unidade {unit_label!r}."
        )

    for (cut_key, _), column in zip(CUTS, columns, strict=True):
        header = _column_header(header_rows, column)
        if cut_key in {"eja_ensino_fundamental", "eja_ensino_medio"}:
            eja_column = columns[9]
            header = f"{_column_header(header_rows, eja_column)} {header}"
        for token in HEADER_TOKENS[cut_key]:
            if token not in header:
                raise ValueError(
                    f"Cabeçalho inconsistente para {cut_key!r} na coluna {column} "
                    f"da aba {sheet_name}: {header!r}."
                )


def _record_from_row(
    row: tuple,
    *,
    row_number: int,
    year: int,
    unit_key: str,
    columns: tuple[int, ...],
    sheet_name: str,
) -> list[dict]:
    municipality_id = _ibge_code(row[3], row_number=row_number)
    municipality_name = _normalise(row[2])
    records = []
    for (cut_key, _), column in zip(CUTS, columns, strict=True):
        records.append({
            "id_municipio": municipality_id,
            "municipio": municipality_name,
            "ano": year,
            "unidade": unit_key,
            "recorte": cut_key,
            "valor": _optional_integer(
                row[column - 1],
                field=f"{unit_key}.{cut_key}",
                row_number=row_number,
            ),
            "tabela_fonte": sheet_name,
            "grupo_comparabilidade": COMPARABILITY_GROUPS.get(
                cut_key,
                "detalhamento_sem_variacao_automatica",
            ),
        })
    return records


def _parse_sheet(workbook, *, year: int, unit_key: str) -> tuple[pd.DataFrame, dict]:
    unit = UNITS[unit_key]
    is_new_layout = year >= 2025
    suffix = unit["new_sheet" if is_new_layout else "old_sheet"]
    columns = unit["new_columns" if is_new_layout else "old_columns"]
    first_data_row = 16 if is_new_layout else 15
    sheet_name = _sheet_name(workbook, suffix)
    sheet = workbook[sheet_name]
    header_rows = list(
        sheet.iter_rows(min_row=1, max_row=first_data_row - 5, values_only=True)
    )
    _validate_layout_headers(
        header_rows,
        unit_key=unit_key,
        columns=columns,
        sheet_name=sheet_name,
    )

    records: list[dict] = []
    municipality_ids: set[str] = set()
    for row_number, row in enumerate(
        sheet.iter_rows(min_row=first_data_row, values_only=True),
        first_data_row,
    ):
        if _normalise(row[1]) != TARGET_UF or not _normalise(row[2]):
            continue
        municipality_id = _ibge_code(row[3], row_number=row_number)
        if municipality_id in municipality_ids:
            raise ValueError(
                f"Código IBGE duplicado na aba {sheet_name}: {municipality_id}."
            )
        municipality_ids.add(municipality_id)
        records.extend(
            _record_from_row(
                row,
                row_number=row_number,
                year=year,
                unit_key=unit_key,
                columns=columns,
                sheet_name=sheet_name,
            )
        )

    if len(municipality_ids) != RS_MUNICIPALITIES:
        raise ValueError(
            f"Cobertura inválida na aba {sheet_name}: esperados {RS_MUNICIPALITIES} "
            f"municípios, encontrados {len(municipality_ids)}."
        )

    return pd.DataFrame(records), {
        "sheet": sheet_name,
        "unit": unit_key,
        "municipalities": len(municipality_ids),
        "rows": len(records),
    }


def _reference_checks(frame: pd.DataFrame) -> dict:
    if TARGET_STATE_CODE != "RS":
        return {}
    checks = {}
    for municipality_id, expected_by_unit in REFERENCE_VALUES_2025.items():
        observed = {}
        for unit_key, expected in expected_by_unit.items():
            values = frame.loc[
                (frame["id_municipio"] == municipality_id)
                & (frame["ano"] == 2025)
                & (frame["unidade"] == unit_key)
                & (frame["recorte"] == "total"),
                "valor",
            ].tolist()
            value = values[0] if len(values) == 1 else None
            observed[unit_key] = value
            if value != expected:
                raise ValueError(
                    f"Validação de referência divergente para {municipality_id}, "
                    f"{unit_key}: esperado {expected}, observado {value}."
                )
        checks[municipality_id] = observed
    return checks


def parse_sources(source_dir: Path, years: tuple[int, ...]) -> tuple[pd.DataFrame, dict]:
    frames = []
    sources = []
    expected_codes: set[str] | None = None
    for year in years:
        if year not in SUPPORTED_YEARS:
            raise ValueError(
                f"Ano {year} não suportado. Use: {', '.join(map(str, SUPPORTED_YEARS))}."
            )
        source_path = _source_path(source_dir, year)
        workbook = load_workbook(source_path, read_only=True, data_only=True)
        try:
            sheet_summaries = []
            year_frames = []
            for unit_key in UNITS:
                unit_frame, sheet_summary = _parse_sheet(
                    workbook,
                    year=year,
                    unit_key=unit_key,
                )
                year_frames.append(unit_frame)
                sheet_summaries.append(sheet_summary)
        finally:
            workbook.close()

        year_frame = pd.concat(year_frames, ignore_index=True)
        year_codes = set(year_frame["id_municipio"].astype(str))
        if expected_codes is None:
            expected_codes = year_codes
        elif year_codes != expected_codes:
            raise ValueError(
                f"A cobertura municipal de {year} difere dos demais anos."
            )
        frames.append(year_frame)
        sources.append({
            "year": year,
            "file": str(source_path),
            "sha256": _file_sha256(source_path),
            "sheets": sheet_summaries,
        })

    frame = pd.concat(frames, ignore_index=True)
    duplicate_columns = ["id_municipio", "ano", "unidade", "recorte"]
    if frame.duplicated(duplicate_columns).any():
        raise ValueError("A tabela longa contém chaves duplicadas.")
    frame["valor"] = pd.array(frame["valor"], dtype="Int64")
    reference_checks = _reference_checks(frame) if 2025 in years else {}
    summary = {
        "years": list(years),
        "municipalities": len(expected_codes or set()),
        "units": list(UNITS),
        "cuts": [cut_key for cut_key, _ in CUTS],
        "rows": len(frame),
        "null_values": int(frame["valor"].isna().sum()),
        "zero_values": int((frame["valor"] == 0).sum()),
        "sources": sources,
        "reference_checks": reference_checks,
    }
    return frame, summary


def replace_years(frame: pd.DataFrame, years: tuple[int, ...]) -> None:
    engine = get_engine("sesi")
    frame = frame.copy()
    frame["sigla_uf"] = TARGET_STATE_CODE
    create_sql = f"""
        CREATE TABLE IF NOT EXISTS {TABLE_NAME} (
            id_municipio VARCHAR(7) NOT NULL,
            municipio TEXT NOT NULL,
            ano INTEGER NOT NULL,
            unidade TEXT NOT NULL,
            recorte TEXT NOT NULL,
            valor INTEGER NULL,
            tabela_fonte TEXT NOT NULL,
            grupo_comparabilidade TEXT NOT NULL,
            sigla_uf VARCHAR(2),
            PRIMARY KEY (id_municipio, ano, unidade, recorte)
        )
    """
    with engine.begin() as connection:
        connection.execute(text(create_sql))
        connection.execute(
            text(
                f"ALTER TABLE {TABLE_NAME} "
                "ADD COLUMN IF NOT EXISTS sigla_uf VARCHAR(2)"
            )
        )
        connection.execute(
            text(
                f"UPDATE {TABLE_NAME} SET sigla_uf = 'RS' "
                "WHERE sigla_uf IS NULL"
            )
        )
        connection.execute(
            text(
                f"DELETE FROM {TABLE_NAME} "
                "WHERE sigla_uf = :state AND ano = ANY(:years)"
            ),
            {"state": TARGET_STATE_CODE, "years": list(years)},
        )
        frame.to_sql(
            TABLE_NAME,
            connection,
            if_exists="append",
            index=False,
            method="multi",
            chunksize=5000,
        )


def main() -> None:
    global RS_MUNICIPALITIES, TARGET_UF, TARGET_STATE_CODE
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--state", default="RS", help="UF da carga (RS ou AL).")
    parser.add_argument(
        "--source-dir",
        type=Path,
        default=SESI_DB_DIR / "data" / "sinopse_estatistica_censo",
        help="Diretório das planilhas oficiais da Sinopse.",
    )
    parser.add_argument(
        "--years",
        type=int,
        nargs="+",
        default=list(SUPPORTED_YEARS),
        help="Anos a validar/importar.",
    )
    parser.add_argument(
        "--apply",
        action="store_true",
        help=f"Substitui os anos selecionados na tabela {TABLE_NAME}.",
    )
    parser.add_argument(
        "--summary-out",
        type=Path,
        help="Grava o resumo auditável em JSON.",
    )
    args = parser.parse_args()

    state_config = load_state_config(args.state)
    TARGET_STATE_CODE = state_config.state_code
    TARGET_UF = state_config.state_name
    RS_MUNICIPALITIES = state_config.expected_municipality_count

    years = tuple(sorted(set(args.years)))
    frame, summary = parse_sources(args.source_dir.resolve(), years)
    summary["state"] = TARGET_STATE_CODE
    if args.apply:
        replace_years(frame, years)
        summary["database_write"] = "applied"
    else:
        summary["database_write"] = "validated_only"

    payload = json.dumps(summary, ensure_ascii=False, indent=2)
    print(payload)
    if args.summary_out:
        args.summary_out.parent.mkdir(parents=True, exist_ok=True)
        args.summary_out.write_text(payload + "\n", encoding="utf-8")


if __name__ == "__main__":
    main()
