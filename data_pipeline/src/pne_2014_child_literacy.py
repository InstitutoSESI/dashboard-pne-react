"""Snapshot canônico do Criança Alfabetizada para o PNE 2014–2024.

O ciclo encerrado usa exclusivamente as planilhas oficiais de divulgação de
2023 e 2024. O resultado estadual vem do agregado oficial do INEP; percentuais
municipais nunca são promediados para construir a referência do RS.
"""

from __future__ import annotations

import hashlib
import json
import math
from functools import lru_cache
from pathlib import Path
from typing import Any, Mapping

import pandas as pd
from openpyxl import load_workbook

from src.child_literacy import (
    EXPECTED_AVAILABLE_BY_STATE,
    EXPECTED_MUNICIPALITIES,
    EXPECTED_STATE_VALUES_BY_STATE,
    MUNICIPAL_UNIVERSE_PATH,
    RS_STATE_CODE,
    RS_STATE_ID,
    _state_source,
    load_municipal_universe,
    parse_state_file,
    sha256_bytes,
    stable_json_bytes,
)
from src.pne_state_context import (
    load_pne_state_context,
    resolve_state_snapshot_dir,
)


CYCLE_ID = "pne_2014_2024"
YEARS = (2023, 2024)
MAX_CYCLE_YEAR = 2024
NETWORK = "municipal"
SOURCE_ID = "inep_avaliacao_alfabetizacao_crianca_alfabetizada"
SOURCE_LABEL = "INEP — Avaliação da Alfabetização / Indicador Criança Alfabetizada."
EXPECTED_AVAILABLE = {2023: 456, 2024: 441}
EXPECTED_STATE_VALUES = {2023: 63.55, 2024: 44.23}
SNAPSHOT_DIR = (
    Path(__file__).resolve().parents[1] / "data" / "pne_2014_child_literacy"
)


def _canonical_ibge_code(value: object) -> str:
    if isinstance(value, bool):
        raise ValueError(f"Código IBGE inválido: {value!r}.")
    if isinstance(value, float):
        if not math.isfinite(value) or not value.is_integer():
            raise ValueError(f"Código IBGE inválido: {value!r}.")
        value = int(value)
    code = str(value or "").strip()
    if not code.isdigit() or len(code) != 7:
        raise ValueError(f"Código IBGE deve ter sete dígitos: {value!r}.")
    return code


def _percentage_or_none(value: object, *, field: str) -> float | None:
    if value is None or str(value).strip() in {"", "-", "—"}:
        return None
    try:
        result = float(str(value).strip().replace(",", "."))
    except ValueError as exc:
        raise ValueError(f"Percentual inválido em {field}: {value!r}.") from exc
    if not math.isfinite(result) or result < 0 or result > 100:
        raise ValueError(f"Percentual fora do domínio em {field}: {value!r}.")
    return result


def _workbook_rows(
    workbook_path: Path,
    *,
    year: int,
    universe: Mapping[str, str],
    reference_date: str,
    state_code: str = "RS",
) -> dict[str, dict[str, Any]]:
    state = load_pne_state_context(state_code)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet_names = [
        name for name in workbook.sheetnames if "divulgação alfabet" in name.casefold()
    ]
    if len(sheet_names) != 1:
        raise ValueError(
            f"Planilha de divulgação não é única em {workbook_path.name}."
        )
    worksheet = workbook[sheet_names[0]]
    header_values = next(
        worksheet.iter_rows(min_row=2, max_row=2, values_only=True)
    )
    headers = [str(value or "").strip() for value in header_values]
    required = {
        "ANO",
        "SG_UF",
        "CO_MUNICIPIO",
        "NO_MUNICIPIO",
        "NO_TP_REDE",
        f"PC_ALUNO_ALFABETIZADO_{year}",
    }
    missing = sorted(required - set(headers))
    if missing:
        raise ValueError(
            f"Colunas obrigatórias ausentes em {workbook_path.name}: {missing}."
        )

    result: dict[str, dict[str, Any]] = {}
    value_field = f"PC_ALUNO_ALFABETIZADO_{year}"
    for values in worksheet.iter_rows(min_row=3, values_only=True):
        row = dict(zip(headers, values))
        if str(row.get("SG_UF") or "").strip().upper() != state.state_code:
            continue
        if str(row.get("NO_TP_REDE") or "").strip().casefold() != NETWORK:
            continue
        if int(row.get("ANO") or 0) != year:
            continue
        municipality_id = _canonical_ibge_code(row.get("CO_MUNICIPIO"))
        if municipality_id not in universe:
            raise ValueError(
                f"Município fora do universo de {state.state_code}: {municipality_id}."
            )
        if municipality_id in result:
            raise ValueError(
                "Duplicidade no grão id_municipio × ano × rede: "
                f"{municipality_id}/{year}/{NETWORK}."
            )
        value = _percentage_or_none(row.get(value_field), field=value_field)
        if value is None:
            continue
        result[municipality_id] = {
            "id_municipio": municipality_id,
            "municipio": universe[municipality_id],
            "ano": year,
            "rede": NETWORK,
            "taxa_alfabetizacao": value,
            "source_id": SOURCE_ID,
            "arquivo_origem": workbook_path.name,
            "data_atualizacao": reference_date,
            "data_status": "available",
        }
    state_expectations = EXPECTED_AVAILABLE_BY_STATE.get(state.state_code)
    if state_expectations is None or year not in state_expectations:
        raise ValueError(
            f"Cobertura oficial esperada ainda não contratada para "
            f"{state.state_code}/{year}."
        )
    expected = state_expectations[year]
    if len(result) != expected:
        raise ValueError(
            f"Cobertura oficial de {year} divergente: {len(result)} != {expected}."
        )
    return result


def _state_rows(
    source_dir: Path,
    *,
    reference_date: str,
    state_code: str = "RS",
) -> tuple[
    list[dict[str, Any]],
    list[dict[str, Any]],
]:
    rows: list[dict[str, Any]] = []
    sources: list[dict[str, Any]] = []
    for year in YEARS:
        content, source_path, member = _state_source(source_dir, year)
        state = load_pne_state_context(state_code)
        parsed = parse_state_file(content, year=year, state_code=state.state_code)
        state_expectations = EXPECTED_STATE_VALUES_BY_STATE.get(state.state_code)
        if state_expectations is None or year not in state_expectations:
            raise ValueError(
                f"Valor estadual esperado ainda não contratado para "
                f"{state.state_code}/{year}."
            )
        expected = state_expectations[year]
        if abs(float(parsed["value"]) - expected) > 1e-9:
            raise ValueError(
                f"Resultado oficial de {state.state_code} em {year} divergente: "
                f"{parsed['value']} != {expected}."
            )
        rows.append(
            {
                "territory_id": state.state_id,
                "territory_name": state.state_name,
                "ano": year,
                "rede": NETWORK,
                "taxa_alfabetizacao": float(parsed["value"]),
                "source_id": SOURCE_ID,
                "arquivo_origem": source_path.name,
                "membro_origem": member,
                "data_atualizacao": reference_date,
                "data_status": "available",
            }
        )
        sources.append(
            {
                "year": year,
                "role": "official_state_municipal_network_percentage",
                "fileName": source_path.name,
                "archiveMember": member,
                "sha256": sha256_bytes(source_path.read_bytes()),
                "memberSha256": sha256_bytes(content),
            }
        )
    return rows, sources


def build_snapshot(
    source_dir: Path,
    *,
    reference_date: str,
    universe_path: Path | None = None,
    state_code: str = "RS",
) -> dict[str, bytes]:
    state = load_pne_state_context(state_code)
    source_root = source_dir.resolve()
    universe = load_municipal_universe(
        universe_path,
        state_code=state.state_code,
    )
    municipal_rows: list[dict[str, Any]] = []
    sources: list[dict[str, Any]] = []

    for year in YEARS:
        workbook_path = source_root / f"alfabetizacao_{year}.xlsx"
        if not workbook_path.is_file():
            raise FileNotFoundError(workbook_path)
        available = _workbook_rows(
            workbook_path,
            year=year,
            universe=universe,
            reference_date=reference_date,
            state_code=state.state_code,
        )
        for municipality_id, municipality_name in sorted(universe.items()):
            municipal_rows.append(
                available.get(municipality_id)
                or {
                    "id_municipio": municipality_id,
                    "municipio": municipality_name,
                    "ano": year,
                    "rede": NETWORK,
                    "taxa_alfabetizacao": None,
                    "source_id": SOURCE_ID,
                    "arquivo_origem": workbook_path.name,
                    "data_atualizacao": reference_date,
                    "data_status": "unavailable",
                }
            )
        sources.append(
            {
                "year": year,
                "role": "official_municipal_network_publication",
                "fileName": workbook_path.name,
                "sha256": sha256_bytes(workbook_path.read_bytes()),
            }
        )

    state_rows, state_sources = _state_rows(
        source_root,
        reference_date=reference_date,
        state_code=state.state_code,
    )
    sources.extend(state_sources)
    municipal_bytes = stable_json_bytes(municipal_rows)
    state_bytes = stable_json_bytes(state_rows)
    manifest = {
        "schemaVersion": "pne-2014-child-literacy-snapshot-v1",
        "stateCode": state.state_code,
        "stateId": state.state_id,
        "stateName": state.state_name,
        "cycle": CYCLE_ID,
        "maximumYear": MAX_CYCLE_YEAR,
        "years": list(YEARS),
        "indicator": "alfabetizacao",
        "network": NETWORK,
        "sourceId": SOURCE_ID,
        "sourceLabel": SOURCE_LABEL,
        "sourceReferenceDate": reference_date,
        "municipalityCount": state.expected_municipality_count,
        "availableByYear": {
            str(year): EXPECTED_AVAILABLE_BY_STATE[state.state_code][year]
            for year in YEARS
        },
        "stateValues": {
            str(year): EXPECTED_STATE_VALUES_BY_STATE[state.state_code][year]
            for year in YEARS
        },
        "canonicalKey": ["id_municipio", "ano", "rede"],
        "sources": sources,
        "files": {
            "municipal_results.json": sha256_bytes(municipal_bytes),
            "state_results.json": sha256_bytes(state_bytes),
        },
    }
    return {
        "municipal_results.json": municipal_bytes,
        "state_results.json": state_bytes,
        "manifest.json": stable_json_bytes(manifest),
    }


def load_snapshot(
    snapshot_dir: Path | None = None,
    *,
    state_code: str | None = None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    state = load_pne_state_context(state_code)
    root = (
        Path(snapshot_dir)
        if snapshot_dir is not None
        else resolve_state_snapshot_dir(SNAPSHOT_DIR, state.state_code)
    ).resolve()
    manifest = json.loads((root / "manifest.json").read_text(encoding="utf-8"))
    if manifest.get("schemaVersion") != "pne-2014-child-literacy-snapshot-v1":
        raise ValueError("Schema do snapshot de alfabetização do PNE 2014 inválido.")
    if manifest.get("stateCode", "RS") != state.state_code:
        raise ValueError("UF do snapshot de alfabetização do PNE 2014 divergente.")
    if int(manifest.get("maximumYear") or 0) != MAX_CYCLE_YEAR:
        raise ValueError("Snapshot do ciclo encerrado permite ano posterior a 2024.")
    for filename, expected_hash in (manifest.get("files") or {}).items():
        if sha256_bytes((root / filename).read_bytes()) != expected_hash:
            raise ValueError(f"Hash divergente no snapshot: {filename}.")
    municipal = json.loads(
        (root / "municipal_results.json").read_text(encoding="utf-8")
    )
    state_results = json.loads(
        (root / "state_results.json").read_text(encoding="utf-8")
    )
    _validate_municipal_rows(
        municipal,
        expected_municipalities=state.expected_municipality_count,
        municipality_ids=state.municipality_ids,
    )
    return municipal, state_results, manifest


def _validate_municipal_rows(
    rows: list[dict[str, Any]],
    *,
    expected_municipalities: int = EXPECTED_MUNICIPALITIES,
    municipality_ids: frozenset[str] | None = None,
) -> None:
    expected_rows = expected_municipalities * len(YEARS)
    if len(rows) != expected_rows:
        raise ValueError(
            f"Snapshot municipal deve conter {expected_rows} linhas canônicas."
        )
    keys: set[tuple[str, int, str]] = set()
    for row in rows:
        municipality_id = _canonical_ibge_code(row.get("id_municipio"))
        year = int(row.get("ano") or 0)
        network = str(row.get("rede") or "")
        if year not in YEARS or year > MAX_CYCLE_YEAR:
            raise ValueError(f"Ano fora do ciclo encerrado: {year}.")
        if network != NETWORK:
            raise ValueError(f"Rede inválida no snapshot: {network!r}.")
        key = (municipality_id, year, network)
        if key in keys:
            raise ValueError(f"Chave canônica duplicada: {key}.")
        keys.add(key)
        value = row.get("taxa_alfabetizacao")
        status = row.get("data_status")
        if value is None:
            if status == "available":
                raise ValueError(f"Resultado disponível sem valor: {key}.")
            continue
        _percentage_or_none(value, field="taxa_alfabetizacao")
    if municipality_ids is not None and {key[0] for key in keys} != municipality_ids:
        raise ValueError("Universo municipal do snapshot de alfabetização divergente.")


@lru_cache(maxsize=4)
def load_dataframe(
    snapshot_dir: Path | None = None,
    *,
    state_code: str | None = None,
) -> pd.DataFrame:
    municipal, _state, _manifest = load_snapshot(
        snapshot_dir,
        state_code=state_code,
    )
    frame = pd.DataFrame(municipal)
    frame["id_municipio"] = frame["id_municipio"].astype("string")
    frame["ano"] = pd.to_numeric(frame["ano"], errors="raise").astype(int)
    frame["taxa_alfabetizacao"] = pd.to_numeric(
        frame["taxa_alfabetizacao"],
        errors="coerce",
    )
    frame["dependencia"] = frame["rede"]
    return frame.sort_values(
        ["id_municipio", "ano", "rede"],
        kind="stable",
    ).reset_index(drop=True)


def snapshot_digest(files: Mapping[str, bytes]) -> str:
    digest = hashlib.sha256()
    for filename, content in sorted(files.items()):
        digest.update(filename.encode("utf-8"))
        digest.update(b"\0")
        digest.update(content)
    return digest.hexdigest()


__all__ = [
    "CYCLE_ID",
    "EXPECTED_AVAILABLE",
    "MAX_CYCLE_YEAR",
    "NETWORK",
    "SNAPSHOT_DIR",
    "SOURCE_ID",
    "SOURCE_LABEL",
    "YEARS",
    "build_snapshot",
    "load_dataframe",
    "load_snapshot",
    "snapshot_digest",
]
