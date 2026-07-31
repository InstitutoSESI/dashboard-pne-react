#!/usr/bin/env python3
"""Incorpora CPC 2023 e Enade Licenciaturas 2025; audita o IGC 2023."""

from __future__ import annotations

import argparse
from collections import Counter, defaultdict
from hashlib import sha256
from pathlib import Path
import ssl
import sys
import tempfile
import urllib.request

import openpyxl


DATA_PIPELINE_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(DATA_PIPELINE_DIR))

from src.pne_macro_ingestion import (  # noqa: E402
    DATA_ROOT,
    MANIFEST_SCHEMA,
    canonical_json_bytes,
    load_municipality_universe,
    normalized_snapshot,
    normalize_name,
    raw_file_entry,
    write_source_snapshot,
)


SOURCE_ID = "inep_quality_offer"
CPC_URL = (
    "https://download.inep.gov.br/educacao_superior/indicadores/"
    "resultados/2023/CPC_2023.xlsx"
)
IGC_URL = (
    "https://download.inep.gov.br/educacao_superior/indicadores/"
    "resultados/2023/IGC_2023.xlsx"
)
ENADE_URL = (
    "https://download.inep.gov.br/educacao_superior/indicadores/"
    "resultados/2025/conceito_enade_licenciaturas.xlsx"
)
SOURCE_PAGE = (
    "https://www.gov.br/inep/pt-br/acesso-a-informacao/dados-abertos/"
    "indicadores-educacionais/indicadores-de-qualidade-da-educacao-superior"
)
DESTINATION = DATA_ROOT / "quality_offer"


def _download(url: str, destination: Path) -> Path:
    request = urllib.request.Request(
        url,
        headers={"User-Agent": "dashboard-pne-data-pipeline/1.0"},
    )
    context = ssl.create_default_context()
    try:
        response = urllib.request.urlopen(request, timeout=180, context=context)
    except Exception:
        context = ssl._create_unverified_context()  # noqa: SLF001
        response = urllib.request.urlopen(request, timeout=180, context=context)
    with response, destination.open("wb") as target:
        while chunk := response.read(1024 * 1024):
            target.write(chunk)
    return destination


def _headers(worksheet: openpyxl.worksheet.worksheet.Worksheet) -> tuple[list[str], dict[str, int]]:
    values = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(value or "").strip() for value in values]
    normalized = {normalize_name(header): index for index, header in enumerate(headers)}
    if len(normalized) != len(headers):
        raise ValueError(f"Cabeçalho duplicado em {worksheet.title}.")
    return headers, normalized


def _position(positions: dict[str, int], *aliases: str) -> int:
    for alias in aliases:
        if normalize_name(alias) in positions:
            return positions[normalize_name(alias)]
    raise ValueError(f"Coluna ausente; alternativas={aliases!r}")


def _integer(value: object) -> int | None:
    if isinstance(value, bool) or value is None:
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if not number.is_integer() or number < 0:
        return None
    return int(number)


def _empty_quality_record(municipality_id: str, name: str) -> dict:
    return {
        "municipalityId": municipality_id,
        "municipalityName": name,
        "cpc2023": {
            "adequateCount": 0,
            "validCount": 0,
            "suppressedCourseCount": 0,
            "organizationAcademic": {},
            "administrativeCategory": {},
        },
        "enadeLicenciaturas2025": {
            "adequateCount": 0,
            "validCount": 0,
            "suppressedCourseCount": 0,
            "suppressedParticipantCount": 0,
            "organizationAcademic": {},
            "administrativeCategory": {},
        },
    }


def _parse_cpc(path: Path, records: dict[str, dict]) -> dict:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[0]
        headers, positions = _headers(worksheet)
        code_index = _position(positions, "Código do Município")
        uf_index = _position(positions, "Sigla da UF")
        course_index = _position(positions, "Código do Curso")
        concept_index = _position(positions, "CPC (Faixa)")
        organization_index = _position(positions, "Organização Acadêmica")
        administration_index = _position(positions, "Categoria Administrativa")
        seen_courses: set[str] = set()
        concept_values: Counter[str] = Counter()
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if str(row[uf_index] or "").strip() != "RS":
                continue
            municipality_id = str(row[code_index] or "").split(".")[0]
            if municipality_id not in records:
                raise ValueError(f"CPC com município RS desconhecido: {municipality_id}")
            course_id = str(row[course_index] or "").strip()
            if not course_id or course_id in seen_courses:
                raise ValueError(f"CPC com curso ausente ou duplicado: {course_id!r}")
            seen_courses.add(course_id)
            concept = _integer(row[concept_index])
            concept_values[str(row[concept_index] or "")] += 1
            component = records[municipality_id]["cpc2023"]
            if concept is None or not 1 <= concept <= 5:
                component["suppressedCourseCount"] += 1
                continue
            component["validCount"] += 1
            component["adequateCount"] += int(concept >= 3)
            for field, index in (
                ("organizationAcademic", organization_index),
                ("administrativeCategory", administration_index),
            ):
                label = str(row[index] or "Não informado").strip()
                component[field][label] = component[field].get(label, 0) + 1
        return {
            "sheet": worksheet.title,
            "headers": headers,
            "rsCourseCount": len(seen_courses),
            "conceptValues": dict(sorted(concept_values.items())),
        }
    finally:
        workbook.close()


def _parse_enade(path: Path, records: dict[str, dict]) -> dict:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[0]
        headers, positions = _headers(worksheet)
        code_index = _position(positions, "Código do Município")
        uf_index = _position(positions, "Sigla da UF")
        course_index = _position(positions, "Código do Curso")
        organization_index = _position(
            positions,
            "Organização Acadêmica ¹",
            "Organização Acadêmica",
        )
        administration_index = _position(
            positions,
            "Categoria Administrativa ²",
            "Categoria Administrativa",
        )
        participants_index = _position(
            positions,
            "Nº  de Concluintes Participantes",
            "Nº de Concluintes Participantes",
        )
        adequate_index = _position(
            positions,
            "Total de Concluinte  Igual ou Acima do Padrão 1 de Proficiência",
        )
        concept_index = _position(positions, "Conceito Enade (Faixa)")
        seen_courses: set[str] = set()
        concept_values: Counter[str] = Counter()
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if str(row[uf_index] or "").strip() != "RS":
                continue
            municipality_id = str(row[code_index] or "").split(".")[0]
            if municipality_id not in records:
                raise ValueError(
                    f"Enade com município RS desconhecido: {municipality_id}"
                )
            course_id = str(row[course_index] or "").strip()
            if not course_id or course_id in seen_courses:
                raise ValueError(f"Enade com curso ausente ou duplicado: {course_id!r}")
            seen_courses.add(course_id)
            participants = _integer(row[participants_index])
            adequate = _integer(row[adequate_index])
            concept_values[str(row[concept_index] or "")] += 1
            component = records[municipality_id]["enadeLicenciaturas2025"]
            if (
                participants is None
                or participants == 0
                or adequate is None
                or adequate > participants
            ):
                component["suppressedCourseCount"] += 1
                component["suppressedParticipantCount"] += participants or 0
                continue
            component["validCount"] += participants
            component["adequateCount"] += adequate
            for field, index in (
                ("organizationAcademic", organization_index),
                ("administrativeCategory", administration_index),
            ):
                label = str(row[index] or "Não informado").strip()
                component[field][label] = component[field].get(label, 0) + 1
        return {
            "sheet": worksheet.title,
            "headers": headers,
            "rsCourseCount": len(seen_courses),
            "conceptValues": dict(sorted(concept_values.items())),
        }
    finally:
        workbook.close()


def _audit_igc(path: Path) -> dict:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[0]
        headers, _ = _headers(worksheet)
        municipality_fields = [
            header for header in headers if "municip" in normalize_name(header)
        ]
        return {
            "sheet": worksheet.title,
            "headers": headers,
            "municipalityFields": municipality_fields,
            "status": "blocked",
            "reason": (
                "O arquivo oficial do IGC 2023 identifica IES e UF, mas não "
                "traz município da sede; atribuir o conceito por município "
                "exigiria uma junção externa não homologada nesta rodada."
            ),
        }
    finally:
        workbook.close()


def parse_quality(cpc_path: Path, enade_path: Path, igc_path: Path) -> tuple[dict, dict]:
    municipality_names, _ = load_municipality_universe()
    records = {
        municipality_id: _empty_quality_record(municipality_id, name)
        for municipality_id, name in municipality_names.items()
    }
    cpc_audit = _parse_cpc(cpc_path, records)
    enade_audit = _parse_enade(enade_path, records)
    igc_audit = _audit_igc(igc_path)
    for record in records.values():
        for component_name in ("cpc2023", "enadeLicenciaturas2025"):
            component = record[component_name]
            component["organizationAcademic"] = dict(
                sorted(component["organizationAcademic"].items())
            )
            component["administrativeCategory"] = dict(
                sorted(component["administrativeCategory"].items())
            )
    normalized = normalized_snapshot(
        source_id=SOURCE_ID,
        edition="CPC 2023; IGC 2023; Enade Licenciaturas 2025",
        records=records,
        municipality_names=municipality_names,
    )
    coverage = {
        "municipalityCount": len(records),
        "cpcMunicipalitiesWithValidResult": sum(
            record["cpc2023"]["validCount"] > 0 for record in records.values()
        ),
        "enadeMunicipalitiesWithValidResult": sum(
            record["enadeLicenciaturas2025"]["validCount"] > 0
            for record in records.values()
        ),
        "cpcValidCourseCount": sum(
            record["cpc2023"]["validCount"] for record in records.values()
        ),
        "enadeValidParticipantCount": sum(
            record["enadeLicenciaturas2025"]["validCount"]
            for record in records.values()
        ),
        "enadeSuppressedParticipantCount": sum(
            record["enadeLicenciaturas2025"]["suppressedParticipantCount"]
            for record in records.values()
        ),
    }
    return normalized, {
        "coverage": coverage,
        "cpc": cpc_audit,
        "enade": enade_audit,
        "igc": igc_audit,
    }


def materialize(cpc_path: Path, enade_path: Path, igc_path: Path) -> dict:
    normalized, audit = parse_quality(cpc_path, enade_path, igc_path)
    specs = [
        ("cpc2023", cpc_path, CPC_URL),
        ("enadeLicenciaturas2025", enade_path, ENADE_URL),
        ("igc2023", igc_path, IGC_URL),
    ]
    manifest = {
        "schemaVersion": MANIFEST_SCHEMA,
        "sourceId": SOURCE_ID,
        "sourceTitle": "Indicadores de Qualidade da Educação Superior",
        "organization": "Instituto Nacional de Estudos e Pesquisas Educacionais Anísio Teixeira",
        "edition": "CPC e IGC 2023; Enade Licenciaturas 2025",
        "sourcePageUrl": SOURCE_PAGE,
        "rawFiles": [
            raw_file_entry(logical_name=name, path=path, official_url=url)
            for name, path, url in specs
        ],
        "dictionary": {
            "cpc": audit["cpc"]["headers"],
            "enadeLicenciaturas": audit["enade"]["headers"],
            "igc": audit["igc"]["headers"],
        },
        "coverage": audit["coverage"],
        "normalization": {
            "municipalityKey": "Código do Município (IBGE, sete dígitos)",
            "cpcAdequate": "curso com CPC (Faixa) 3, 4 ou 5",
            "enadeAdequate": (
                "Total de Concluinte Igual ou Acima do Padrão 1 de Proficiência"
            ),
            "suppression": "SC, vazio ou componente inconsistente é unknown",
            "normalizedSha256": sha256(canonical_json_bytes(normalized)).hexdigest(),
        },
        "absencePolicy": {
            "noLocalHigherEducationOffer": "not_applicable",
            "offerWithoutEvaluationInCycle": "unavailable",
            "suppressedResult": "unknown; never zero",
        },
        "duplicatePolicy": "código de curso repetido no ciclo invalida a carga",
        "cycles": {
            "cpc": "2023; não interpolado como série anual",
            "enadeLicenciaturas": "2025; não interpolado como série anual",
        },
        "preservedDimensions": [
            "Organização Acadêmica",
            "Categoria Administrativa",
        ],
        "blocked": {"igcMunicipalization": audit["igc"]},
        "status": "partially_approved_complementary",
        "approvedComponents": ["cpc2023", "enadeLicenciaturas2025"],
    }
    write_source_snapshot(
        destination=DESTINATION,
        raw_files={path.name: path for _, path, _ in specs},
        normalized=normalized,
        manifest=manifest,
    )
    return manifest


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--cpc-file", type=Path)
    parser.add_argument("--enade-file", type=Path)
    parser.add_argument("--igc-file", type=Path)
    args = parser.parse_args()
    supplied = (args.cpc_file, args.enade_file, args.igc_file)
    if any(supplied) and not all(supplied):
        parser.error("Informe os três arquivos locais ou nenhum.")
    with tempfile.TemporaryDirectory(prefix="pne-quality-") as temporary:
        root = Path(temporary)
        files = supplied if all(supplied) else (
            _download(CPC_URL, root / "CPC_2023.xlsx"),
            _download(ENADE_URL, root / "conceito_enade_licenciaturas.xlsx"),
            _download(IGC_URL, root / "IGC_2023.xlsx"),
        )
        manifest = materialize(*files)
    print(canonical_json_bytes(manifest).decode("utf-8"), end="")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
