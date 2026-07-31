"""Indicador 11.d: atendimento EJA 18+ em base territorial mista."""

from __future__ import annotations

import hashlib
import json
import math
from pathlib import Path
from typing import Any, Mapping

from openpyxl import load_workbook


YEARS = (2024, 2025)
SHEETS = {2024: "1.38", 2025: "1.53"}
EXPECTED_MUNICIPALITIES = 497
EXPECTED_ZERO_NUMERATORS = {2024: 247, 2025: 249}
EXPECTED_RS_VALUES = {
    2024: 1.5272312091296365,
    2025: 1.3658147966780407,
}

DATA_DIR = Path(__file__).resolve().parents[1] / "data"
SNAPSHOT_DIR = DATA_DIR / "pne_goal_11d_eja"
DENOMINATOR_DIR = DATA_DIR / "pne_goal_11b_census_2022"


def stable_json_bytes(value: Any) -> bytes:
    return (
        json.dumps(
            value,
            ensure_ascii=False,
            indent=2,
            sort_keys=True,
            allow_nan=False,
        )
        + "\n"
    ).encode("utf-8")


def sha256_bytes(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def _integer(value: object, *, context: str) -> int:
    if value in (None, "", "-", "..", "..."):
        raise ValueError(f"Valor ausente em {context}.")
    numeric = float(value)
    if not math.isfinite(numeric) or numeric < 0 or not numeric.is_integer():
        raise ValueError(f"Contagem inválida em {context}: {value!r}.")
    return int(numeric)


def load_denominators(
    denominator_dir: Path = DENOMINATOR_DIR,
) -> tuple[dict[str, dict[str, Any]], dict[str, Any]]:
    manifest_path = denominator_dir / "manifest.json"
    components_path = denominator_dir / "municipal_components.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    expected_hash = (manifest.get("files") or {}).get("municipal_components.json")
    if expected_hash != sha256_bytes(components_path.read_bytes()):
        raise ValueError("Hash do denominador censitário 11.d divergente.")
    rows = json.loads(components_path.read_text(encoding="utf-8"))
    result = {}
    for row in rows:
        source = row["eighteenPlus"]["sourceValues"]
        denominator = int(source["below_fundamental"]["value"]) + int(
            source["fundamental_complete"]["value"]
        )
        municipality_id = str(row["municipalityId"])
        if municipality_id in result:
            raise ValueError(f"Denominador duplicado: {municipality_id}.")
        result[municipality_id] = {
            "municipalityId": municipality_id,
            "municipalityName": str(row["municipalityName"]),
            "denominator": denominator,
        }
    if len(result) != EXPECTED_MUNICIPALITIES:
        raise ValueError("O denominador 11.d não cobre os 497 municípios.")
    return result, manifest


def _workbook_for_year(source_dir: Path, year: int) -> Path:
    matches = sorted(source_dir.glob(f"*{year}.xlsx"))
    if len(matches) != 1:
        raise FileNotFoundError(
            f"Esperado um workbook oficial único da Sinopse {year}."
        )
    return matches[0]


def parse_eja_workbook(path: Path, *, year: int) -> dict[str, int]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_marker = SHEETS[year]
        matches = [name for name in workbook.sheetnames if sheet_marker in name]
        if len(matches) != 1:
            raise ValueError(
                f"Tabela {sheet_marker} não é única na Sinopse {year}."
            )
        worksheet = workbook[matches[0]]
        title = str(worksheet.cell(4, 1).value or "")
        if "Educação de Jovens e Adultos" not in title or str(year) not in title:
            raise ValueError(f"Título semântico inesperado na tabela de {year}.")
        age_headers = [
            str(worksheet.cell(8, column).value or "").strip()
            for column in range(8, 14)
        ]
        expected_headers = [
            "18 a 19 anos",
            "20 a 24 anos",
            "25 a 29 anos",
            "30 a 34 anos",
            "35 a 39 anos",
            "40 anos ou mais",
        ]
        if age_headers != expected_headers:
            raise ValueError(f"Faixas etárias inesperadas na Sinopse {year}.")

        rows: dict[str, int] = {}
        for index, row in enumerate(
            worksheet.iter_rows(min_row=10, values_only=True),
            start=10,
        ):
            if str(row[1] or "").strip() != "Rio Grande do Sul":
                continue
            municipality_name = str(row[2] or "").strip()
            municipality_id = str(row[3] or "").strip().split(".")[0]
            if not municipality_name or not municipality_id:
                continue
            if municipality_id in rows:
                raise ValueError(
                    f"Município duplicado na Sinopse {year}: {municipality_id}."
                )
            rows[municipality_id] = sum(
                _integer(row[column], context=f"{year}/linha {index}/col {column+1}")
                for column in range(7, 13)
            )
        if len(rows) != EXPECTED_MUNICIPALITIES:
            raise ValueError(
                f"Cobertura municipal EJA {year}: {len(rows)} != 497."
            )
        zero_count = sum(value == 0 for value in rows.values())
        if zero_count != EXPECTED_ZERO_NUMERATORS[year]:
            raise ValueError(
                f"Zeros explícitos EJA {year}: {zero_count} "
                f"!= {EXPECTED_ZERO_NUMERATORS[year]}."
            )
        return rows
    finally:
        workbook.close()


def ratio(numerator: int, denominator: int) -> dict[str, Any]:
    if denominator == 0:
        return {
            "dataStatus": "not_applicable",
            "reasonCode": "denominator_zero",
            "value": None,
            "numerator": numerator,
            "denominator": denominator,
        }
    value = 100.0 * numerator / denominator
    if not math.isfinite(value) or value < 0:
        raise ValueError("Razão EJA inválida.")
    return {
        "dataStatus": "available",
        "value": value,
        "numerator": numerator,
        "denominator": denominator,
    }


def build_snapshot(
    source_dir: Path,
    *,
    reference_date: str,
    denominator_dir: Path = DENOMINATOR_DIR,
) -> dict[str, bytes]:
    denominators, denominator_manifest = load_denominators(denominator_dir)
    by_year: dict[int, dict[str, int]] = {}
    sources = []
    for year in YEARS:
        path = _workbook_for_year(source_dir.resolve(), year)
        by_year[year] = parse_eja_workbook(path, year=year)
        sources.append(
            {
                "year": year,
                "table": SHEETS[year],
                "fileName": path.name,
                "sha256": sha256_bytes(path.read_bytes()),
            }
        )

    municipal = []
    for municipality_id, base in sorted(denominators.items()):
        series = []
        for year in YEARS:
            result = ratio(
                by_year[year][municipality_id],
                int(base["denominator"]),
            )
            series.append(
                {
                    "municipalityId": municipality_id,
                    "municipalityName": base["municipalityName"],
                    "year": year,
                    **result,
                }
            )
        municipal.append(
            {
                "municipalityId": municipality_id,
                "municipalityName": base["municipalityName"],
                "series": series,
            }
        )

    state = []
    denominator_sum = sum(
        int(item["denominator"]) for item in denominators.values()
    )
    for year in YEARS:
        numerator_sum = sum(by_year[year].values())
        value = 100.0 * numerator_sum / denominator_sum
        if abs(value - EXPECTED_RS_VALUES[year]) > 1e-12:
            raise ValueError(f"Reconciliação estadual 11.d divergente em {year}.")
        state.append(
            {
                "territoryId": "43",
                "territoryName": "Rio Grande do Sul",
                "year": year,
                "dataStatus": "available",
                "value": value,
                "numerator": numerator_sum,
                "denominator": denominator_sum,
            }
        )

    municipal_bytes = stable_json_bytes(municipal)
    state_bytes = stable_json_bytes(state)
    manifest = {
        "schemaVersion": "pne-goal-11d-eja-snapshot-v1",
        "sourceReferenceDate": reference_date,
        "indicator": "eja_atendimento_18_mais",
        "territorialBasis": {
            "numerator": "municipality_of_school",
            "denominator": "municipality_of_residence",
        },
        "municipalityCount": EXPECTED_MUNICIPALITIES,
        "years": list(YEARS),
        "zeroNumeratorsByYear": {
            str(year): EXPECTED_ZERO_NUMERATORS[year] for year in YEARS
        },
        "stateValues": {
            str(year): EXPECTED_RS_VALUES[year] for year in YEARS
        },
        "sources": sources,
        "denominatorSnapshot": {
            "schemaVersion": denominator_manifest["schemaVersion"],
            "manifestSha256": sha256_bytes(
                (denominator_dir / "manifest.json").read_bytes()
            ),
        },
        "files": {
            "municipal_results.json": sha256_bytes(municipal_bytes),
            "state_results.json": sha256_bytes(state_bytes),
        },
    }
    return {
        "municipal_results.json": municipal_bytes,
        "state_results.json": state_bytes,
        "manifest.json": stable_json_bytes(manifest),
    }


def load_snapshot(
    snapshot_dir: Path = SNAPSHOT_DIR,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    root = snapshot_dir.resolve()
    manifest = json.loads((root / "manifest.json").read_text(encoding="utf-8"))
    if manifest.get("schemaVersion") != "pne-goal-11d-eja-snapshot-v1":
        raise ValueError("Schema do snapshot 11.d inválido.")
    for filename, expected in (manifest.get("files") or {}).items():
        if sha256_bytes((root / filename).read_bytes()) != expected:
            raise ValueError(f"Hash divergente no snapshot 11.d: {filename}.")
    municipal = json.loads(
        (root / "municipal_results.json").read_text(encoding="utf-8")
    )
    state = json.loads((root / "state_results.json").read_text(encoding="utf-8"))
    if len(municipal) != EXPECTED_MUNICIPALITIES:
        raise ValueError("Cobertura municipal do snapshot 11.d inválida.")
    return municipal, state, manifest


def current_results(
    rows: list[Mapping[str, Any]],
    *,
    year: int = 2025,
) -> dict[str, dict[str, Any]]:
    output = {}
    for row in rows:
        matches = [item for item in row["series"] if int(item["year"]) == year]
        if len(matches) != 1:
            raise ValueError(f"Série 11.d sem ano único: {row['municipalityId']}.")
        output[str(row["municipalityId"])] = dict(matches[0])
    return output
